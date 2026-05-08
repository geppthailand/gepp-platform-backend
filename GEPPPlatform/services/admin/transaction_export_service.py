"""
Admin Transaction Export Service — XLSX export of v3 transactions for an
organization. Column layout mirrors the GEPP-Business v2 "Export XLSX"
output exactly: 13 columns, one row per transaction record, with the
Transaction ID formatted as "{tx.id}-{n}" where n is the record's 1-based
position in `transactions.transaction_records` (the on-DB array order).

Status is derived from the records (mirrors ManualAuditService):
  - all records 'approved' → 'approved'
  - any record 'rejected'  → 'rejected'
  - else                   → 'pending'
"""

import base64
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from GEPPPlatform.exceptions import BadRequestException, NotFoundException
from GEPPPlatform.models.subscriptions.organizations import (
    Organization,
    OrganizationSetup,
)
from GEPPPlatform.models.transactions.transaction_records import TransactionRecord
from GEPPPlatform.models.transactions.transactions import Transaction
from GEPPPlatform.models.users.user_location import UserLocation
from GEPPPlatform.models.users.user_related import UserLocationTag


# Default level labels when the org hasn't customised them.
DEFAULT_LEVEL_LABELS = ('Branch', 'Building', 'Floor', 'Room')

# All user-facing dates and date filters are interpreted in this zone.
# `transactions.transaction_date` is `timestamp with time zone` stored in
# UTC; we convert to Bangkok at the boundary so the user's calendar
# view matches what they typed and what gets exported.
DISPLAY_TZ = ZoneInfo('Asia/Bangkok')


# ── Helpers ────────────────────────────────────────────────────────────────

def _derive_status(record_statuses: List[Optional[str]]) -> str:
    if not record_statuses:
        return 'pending'
    statuses = {(s or 'pending') for s in record_statuses}
    if statuses == {'approved'}:
        return 'approved'
    if 'rejected' in statuses:
        return 'rejected'
    return 'pending'


def _parse_date_from(value: Optional[str]) -> Optional[datetime]:
    """Parse the lower bound. A plain `YYYY-MM-DD` is anchored to
    00:00:00 in DISPLAY_TZ (Asia/Bangkok); a full ISO timestamp is
    accepted as-is. Returns a tz-aware datetime so SQLAlchemy can
    compare it against `timestamp with time zone` columns directly."""
    if not value:
        return None
    try:
        if 'T' not in value:
            d = datetime.strptime(value, '%Y-%m-%d')
            return d.replace(tzinfo=DISPLAY_TZ)
        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=DISPLAY_TZ)
        return dt
    except ValueError:
        raise BadRequestException(f'Invalid date format: {value}')


def _parse_date_to(value: Optional[str]) -> Optional[datetime]:
    """Parse the upper bound. A plain `YYYY-MM-DD` snaps to
    23:59:59.999999 in DISPLAY_TZ so the user's "to 30/04" filter
    actually includes April 30 transactions in their local calendar."""
    if not value:
        return None
    try:
        if 'T' not in value:
            d = datetime.strptime(value, '%Y-%m-%d')
            return d.replace(
                hour=23, minute=59, second=59, microsecond=999999,
                tzinfo=DISPLAY_TZ,
            )
        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=DISPLAY_TZ)
        return dt
    except ValueError:
        raise BadRequestException(f'Invalid date format: {value}')


def _fmt_bkk_date(dt: Optional[datetime]) -> str:
    """Render a tz-aware datetime as DD/MM/YYYY in DISPLAY_TZ. Naive
    inputs are assumed to already be in Bangkok local time."""
    if dt is None:
        return ''
    if dt.tzinfo is not None:
        dt = dt.astimezone(DISPLAY_TZ)
    return dt.strftime('%d/%m/%Y')


def _loc_label(loc: Optional[UserLocation]) -> str:
    if not loc:
        return ''
    return loc.display_name or loc.name_en or loc.name_th or f'#{loc.id}'


# ── Service ───────────────────────────────────────────────────────────────

class AdminTransactionExportService:
    """Builds a v2-shaped XLSX server-side and returns it base64-encoded."""

    # Static portion of the header. The 4 hierarchical "Location" columns
    # are inserted dynamically based on the org's level-name substitution,
    # so the actual headers are built in `_resolve_level_labels()`.
    HEADERS_BEFORE_LOCATION = ['#', 'Transaction Date', 'Transaction ID']
    HEADERS_AFTER_LOCATION = [
        'Location Tag', 'Destination',
        'Main Material', 'Sub Material',
        'Weight (Kg)', 'Price per Kg', 'Total Price (THB)',
        'Status', 'Note',
    ]

    # Sort fields exposed via the `sort` query param. Each maps to a
    # SQLAlchemy column on Transaction. Direction is appended as ":asc" or
    # ":desc" — e.g. "transaction_date:asc".
    SORT_FIELDS = {
        'transaction_date': Transaction.transaction_date,
        'created_date': Transaction.created_date,
        'id': Transaction.id,
        'weight_kg': Transaction.weight_kg,
        'total_amount': Transaction.total_amount,
    }

    def __init__(self, db_session: Session):
        self.db = db_session

    # ── Public ────────────────────────────────────────────────────────
    def export(self, organization_id: int, query_params: dict) -> Dict[str, Any]:
        org = (
            self.db.query(Organization)
            .filter(Organization.id == organization_id)
            .first()
        )
        if not org:
            raise NotFoundException(f'Organization {organization_id} not found')

        origin_id = query_params.get('originId')
        date_from = _parse_date_from(query_params.get('dateFrom'))
        date_to = _parse_date_to(query_params.get('dateTo'))
        status_filter = (query_params.get('status') or 'all').strip().lower()
        if status_filter not in ('all', 'pending', 'approved', 'rejected'):
            raise BadRequestException(
                f"Invalid status '{status_filter}'. Use all|pending|approved|rejected.")

        sort_raw = (query_params.get('sort') or 'transaction_date:desc').strip().lower()
        sort_field, _, sort_dir = sort_raw.partition(':')
        sort_dir = sort_dir or 'desc'
        if sort_field not in self.SORT_FIELDS:
            raise BadRequestException(
                f"Invalid sort field '{sort_field}'. Allowed: {', '.join(self.SORT_FIELDS)}.")
        if sort_dir not in ('asc', 'desc'):
            raise BadRequestException(
                f"Invalid sort direction '{sort_dir}'. Allowed: asc, desc.")

        level_labels, path_by_node_id = self._resolve_org_setup(organization_id)

        rows = self._collect_rows(
            organization_id=organization_id,
            origin_id=int(origin_id) if origin_id else None,
            date_from=date_from,
            date_to=date_to,
            status_filter=status_filter,
            sort_field=sort_field,
            sort_dir=sort_dir,
            path_by_node_id=path_by_node_id,
        )

        wb = self._build_workbook(rows, level_labels)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode('utf-8')

        ts = datetime.now().strftime('%Y-%m-%d %H_%M_%S')
        filename = f'Export_GEPP_Business_Transaction_{ts}.xlsx'

        return {
            'filename': filename,
            'rowCount': sum(1 for _ in rows),
            'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'base64': b64,
        }

    # ── Level labels & location path resolution ───────────────────────
    def _resolve_org_setup(
        self, organization_id: int
    ) -> 'tuple[List[str], Dict[int, List[int]]]':
        """Returns (column_labels, nodeId→ancestor-path map).

        The label list is the 4 column headers for the hierarchical Location
        block (substitution names where set, defaults otherwise).

        The path map is built by DFS over `OrganizationSetup.root_nodes`.
        Each value is the full chain of nodeIds from a depth-0 root entry
        down to the node itself — its **position in `root_nodes`'s nested
        JSON**, not the materialised `user_locations.organization_path`
        (which is unset for orgs whose tree is built through the org-setup
        flow). Position in the path determines the column:
            index 0 = top-level entry in root_nodes → Branch
            index 1 = its child                     → Building
            index 2 = its grandchild                → Floor
            index 3                                 → Room

        Levels deeper than 3 are ignored. Same applies to nodes appearing
        only inside `hub_node` — those aren't part of the branch/building/
        floor/room taxonomy."""
        setup = (
            self.db.query(OrganizationSetup)
            .filter(
                OrganizationSetup.organization_id == organization_id,
                OrganizationSetup.is_active == True,  # noqa: E712
            )
            .order_by(OrganizationSetup.created_date.desc())
            .first()
        )
        if not setup:
            setup = (
                self.db.query(OrganizationSetup)
                .filter(OrganizationSetup.organization_id == organization_id)
                .order_by(OrganizationSetup.created_date.desc())
                .first()
            )

        labels = list(DEFAULT_LEVEL_LABELS)
        path_by_id: Dict[int, List[int]] = {}

        if setup:
            for idx, attr in enumerate((
                'branch_level_name',
                'building_level_name',
                'floor_level_name',
                'room_level_name',
            )):
                value = getattr(setup, attr, None)
                if value and str(value).strip():
                    labels[idx] = str(value).strip()

            self._index_node_paths(setup.root_nodes or [], current_path=[], path_by_id=path_by_id)

        return labels, path_by_id

    @classmethod
    def _index_node_paths(
        cls,
        nodes: List[Dict[str, Any]],
        current_path: List[int],
        path_by_id: Dict[int, List[int]],
    ) -> None:
        """DFS that records the full ancestor path for each node."""
        if not nodes:
            return
        for node in nodes:
            if not isinstance(node, dict):
                continue
            nid_raw = node.get('nodeId')
            new_path = list(current_path)
            try:
                nid = int(nid_raw) if nid_raw is not None else None
            except (ValueError, TypeError):
                nid = None
            if nid is not None:
                new_path.append(nid)
                # First-write-wins: nodes should appear once; be defensive.
                path_by_id.setdefault(nid, new_path)
            cls._index_node_paths(node.get('children') or [], new_path, path_by_id)

    def _build_location_lookup(self, location_ids: List[int]) -> Dict[int, str]:
        """Batch-resolve location IDs → display label."""
        if not location_ids:
            return {}
        rows = (
            self.db.query(
                UserLocation.id,
                UserLocation.display_name,
                UserLocation.name_en,
                UserLocation.name_th,
            )
            .filter(UserLocation.id.in_(set(location_ids)))
            .all()
        )
        out: Dict[int, str] = {}
        for r in rows:
            out[r.id] = r.display_name or r.name_en or r.name_th or f'#{r.id}'
        return out

    # ── Query ─────────────────────────────────────────────────────────
    def _collect_rows(
        self,
        organization_id: int,
        origin_id: Optional[int],
        date_from: Optional[datetime],
        date_to: Optional[datetime],
        status_filter: str,
        sort_field: str,
        sort_dir: str,
        path_by_node_id: Dict[int, List[int]],
    ) -> list:
        q = (
            self.db.query(Transaction)
            .options(joinedload(Transaction.origin))
            .filter(Transaction.organization_id == organization_id)
            .filter(Transaction.is_active == True)  # noqa: E712
            .filter(Transaction.deleted_date.is_(None))
        )
        if origin_id is not None:
            q = q.filter(Transaction.origin_id == origin_id)
        if date_from is not None:
            q = q.filter(Transaction.transaction_date >= date_from)
        if date_to is not None:
            q = q.filter(Transaction.transaction_date <= date_to)

        sort_col = self.SORT_FIELDS[sort_field]
        order_clause = sort_col.asc() if sort_dir == 'asc' else sort_col.desc()
        # Stable secondary order on id so paged exports / tied dates are
        # deterministic — matches v2 behaviour.
        secondary = Transaction.id.asc() if sort_dir == 'asc' else Transaction.id.desc()
        transactions: List[Transaction] = (
            q.order_by(order_clause, secondary).all()
        )
        if not transactions:
            return []

        # Resolve location-tag names in one shot.
        tag_ids = {t.location_tag_id for t in transactions if t.location_tag_id}
        tag_name_by_id: Dict[int, str] = {}
        if tag_ids:
            for row in (
                self.db.query(UserLocationTag.id, UserLocationTag.name)
                .filter(UserLocationTag.id.in_(tag_ids))
                .all()
            ):
                tag_name_by_id[row.id] = row.name

        # Pull all candidate records.
        all_record_ids: List[int] = []
        for t in transactions:
            for rid in (t.transaction_records or []):
                all_record_ids.append(rid)

        records_by_id: Dict[int, TransactionRecord] = {}
        if all_record_ids:
            recs = (
                self.db.query(TransactionRecord)
                .options(
                    joinedload(TransactionRecord.material),
                    joinedload(TransactionRecord.main_material),
                    joinedload(TransactionRecord.destination),
                    # Currency intentionally NOT eager-loaded — prod schema
                    # is missing the model's `name` column.
                )
                .filter(TransactionRecord.id.in_(all_record_ids))
                .filter(TransactionRecord.is_active == True)  # noqa: E712
                .filter(TransactionRecord.deleted_date.is_(None))
                .all()
            )
            for r in recs:
                records_by_id[r.id] = r

        # Resolve every origin's hierarchical path through the org chart
        # by looking up its ancestor chain in `path_by_node_id` — built
        # from `root_nodes` JSON nesting. The position of each node in
        # that chain (0..3) selects the column (Branch/Building/Floor/Room).
        # Origins missing from the tree (orphan / hub-only / legacy) just
        # land in the Branch column as a best-effort label.
        all_lookup_ids: set = set()
        for tx in transactions:
            if tx.origin_id and tx.origin_id in path_by_node_id:
                all_lookup_ids.update(path_by_node_id[tx.origin_id])
            elif tx.origin_id:
                all_lookup_ids.add(tx.origin_id)

        location_lookup = self._build_location_lookup(list(all_lookup_ids))

        out = []
        for tx in transactions:
            ordered_ids = list(tx.transaction_records or [])
            tx_records = [records_by_id[rid] for rid in ordered_ids if rid in records_by_id]
            derived = _derive_status([r.status for r in tx_records])
            if status_filter != 'all' and derived != status_filter:
                continue
            tag_label = tag_name_by_id.get(tx.location_tag_id, '') if tx.location_tag_id else ''

            level_columns: List[str] = ['', '', '', '']
            if tx.origin_id:
                ancestor_ids = path_by_node_id.get(tx.origin_id)
                if ancestor_ids:
                    for col_idx, pid in enumerate(ancestor_ids[:4]):
                        level_columns[col_idx] = location_lookup.get(pid, f'#{pid}')
                else:
                    # Origin not in the tree — show the location's own
                    # name in Branch as a fallback so the row isn't blank.
                    level_columns[0] = location_lookup.get(
                        tx.origin_id, _loc_label(tx.origin)
                    )

            out.append((tx, tx_records, derived, tag_label, level_columns))
        return out

    # ── Workbook ──────────────────────────────────────────────────────
    def _build_workbook(self, rows: list, level_labels: List[str]) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # Dynamic header: the 4 hierarchical Location columns sit between
        # "Transaction ID" and "Location Tag", using the org's substitution
        # labels (e.g. "สาขา / อาคาร / ชั้น / ห้อง"). No combined "Location"
        # / "Location Path" column — the 4 columns *are* the location.
        headers = (
            list(self.HEADERS_BEFORE_LOCATION)
            + list(level_labels)
            + list(self.HEADERS_AFTER_LOCATION)
        )

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2F855A', end_color='2F855A', fill_type='solid')
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx = 2
        seq = 0
        for tx, tx_records, derived, tag_label, level_columns in rows:
            tx_date_str = _fmt_bkk_date(tx.transaction_date)

            # `level_columns` is already in [branch, building, floor, room]
            # order, with empty strings for levels the origin's path skips.
            level_values = list(level_columns)

            for index, rec in enumerate(tx_records, start=1):
                seq += 1
                tx_id_label = f'{tx.id}-{index}'
                dest_label = _loc_label(rec.destination)

                main_mat = ''
                if rec.main_material:
                    main_mat = rec.main_material.name_en or rec.main_material.name_th or ''
                sub_mat = ''
                if rec.material:
                    sub_mat = rec.material.name_en or rec.material.name_th or ''

                weight = float(rec.origin_weight_kg or 0)
                price_per_kg = float(rec.origin_price_per_unit or 0)
                total_price = float(rec.total_amount or 0)

                values = (
                    [seq, tx_date_str, tx_id_label]
                    + level_values
                    + [
                        tag_label,
                        dest_label,
                        main_mat,
                        sub_mat,
                        weight,
                        price_per_kg,
                        total_price,
                        derived,
                        rec.notes or tx.notes or '',
                    ]
                )
                for col_idx, val in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                row_idx += 1

        # Auto-size columns (cap at 50)
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = 8
            for cell in col:
                v = '' if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

        ws.freeze_panes = 'A2'
        return wb
