"""
Import Organization Setup — back-office service.

One `organization_setup_imports` row per uploaded 5-tab xlsx. Flow mirrors the transaction
import (import_files/ImportService):
  upload   → store raw xlsx in S3 + create batch row (status='uploaded')
  extract  → parse + validate + build an editable preview (status='extracted')
  confirm  → create users/tags/tenants (INSERT) then origins/destinations (REPLACE: a new
             organization_setup version whose tree holds only the imported nodes; the previous
             tree's nodes stay active → become recycle-bin orphans). Records every created id on
             the batch so it can be reverted. (status='confirmed')
  revert   → soft-delete the created users/tags/tenants/locations + insert a new setup version
             with the imported nodeIds stripped out. (status='reverted')

Semantics (user-confirmed):
  - Email uniqueness is GLOBAL (whole system) — duplicates are blocking, fixed in the preview.
  - origins/destinations REPLACE the whole chart; users/tags/tenants INSERT (append).
  - revert only removes what the import created (no snapshot restore of the previous chart).

Composes existing creation logic: UserService.create_user, LocationTagService.create_tag,
TenantService.create_tenant. Origin/destination UserLocation rows are created directly (so tags/
tenants/materials/business_type + the node-level is_destination flag are all set — create_organization_
setup does not carry those) and the OrganizationSetup version is inserted the same way core does.
"""

from __future__ import annotations

import logging
import re
import secrets
import string
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from ....models.organization_setup_import import OrganizationSetupImport
from ....models.subscriptions.organizations import OrganizationSetup, Organization
from ....models.users.user_location import UserLocation
from ....models.users.user_related import UserLocationTag, UserTenant
from ....models.cores.references import Material
from ...file_upload_service import S3FileUploadService
from . import parser as P

logger = logging.getLogger(__name__)

# xlsx Role label → OrganizationRole.key
_ROLE_MAP = {
    'admin': 'admin', 'administrator': 'admin',
    'datainputer': 'data_input', 'datainput': 'data_input', 'datainputspecialist': 'data_input',
    'dataauditor': 'auditor', 'auditor': 'auditor',
    'viewer': 'viewer',
}
# OrganizationRole.key → xlsx Role label (for export). Inverse of the labels _ROLE_MAP accepts.
_ROLE_LABEL = {
    'admin': 'Admin', 'data_input': 'Data Inputer', 'auditor': 'Data Auditor', 'viewer': 'Viewer',
}
# xlsx Business Type (Thai) → hub_type key. Unknown values fall back to the raw string
# (still non-null so the row is recognised as a destination: hub_type IS NOT NULL).
_HUB_TYPE_MAP = {
    'ผู้รวบรวม': 'Collectors',
    'ผู้คัดแยก': 'Sorters',
    'ผู้รวบรวมและคัดแยก': 'Aggregators',
    'สถานีขนถ่าย': 'Transfer Station',
    'อัดก้อน': 'Baler',
    'mrf': 'MRF',
    'โรงงานรีไซเคิล': 'Recycling Plant',
    'หลุมฝังกลบ': 'Landfill',
    'เตาเผาร่วม': 'Co-processing',
    'เตาเผา': 'Incinerator',
}

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def _norm(s: Any) -> str:
    return re.sub(r'\s+', '', str(s or '').strip().lower())


def _role_key(role_label: str) -> Optional[str]:
    return _ROLE_MAP.get(_norm(role_label))


def _hub_type_key(business_type: str) -> str:
    return _HUB_TYPE_MAP.get(_norm(business_type)) or (business_type or '').strip() or 'Collectors'


def _gen_password(length: int = 10) -> str:
    """Readable random password for users imported without one (letters+digits, no ambigu/symbols)."""
    alphabet = ''.join(c for c in (string.ascii_letters + string.digits) if c not in 'Il1O0')
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def _to_dt(s: Any) -> Optional[datetime]:
    """ISO 'YYYY-MM-DD' string → datetime (for the upsert path, which sets columns directly)."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], '%Y-%m-%d')
    except (TypeError, ValueError):
        return None


class SetupImportService:
    def __init__(self, db):
        self.db = db

    # ── 1. Upload ────────────────────────────────────────────────────────────────
    def upload(self, organization_id: int, admin_id: Optional[int], filename: str,
               file_bytes: bytes, content_type: Optional[str] = None) -> Dict[str, Any]:
        try:
            s3 = S3FileUploadService()
            uploaded = s3.upload_import_file(
                file_data=file_bytes, filename=filename, content_type=content_type,
                import_type='organization_setup', organization_id=organization_id,
            )
            row = OrganizationSetupImport(
                organization_id=organization_id, uploaded_by_id=admin_id,
                original_filename=filename,
                s3_key=(uploaded or {}).get('s3_key'), s3_bucket=(uploaded or {}).get('s3_bucket'),
                file_size=(uploaded or {}).get('file_size'),
                mime_type=(uploaded or {}).get('content_type') or content_type,
                status='uploaded',
            )
            self.db.add(row)
            self.db.commit()
            self.db.refresh(row)
            return {'success': True, 'data': self._row_meta(row)}
        except Exception as e:
            self.db.rollback()
            logger.error(f"setup-import upload error: {e}")
            return {'success': False, 'message': 'Upload failed', 'errors': [str(e)]}

    # ── 2. Extract + validate → preview ──────────────────────────────────────────
    def extract(self, import_id: int, organization_id: int) -> Dict[str, Any]:
        row = self._get_row(import_id, organization_id)
        if not row:
            return {'success': False, 'message': 'Import not found'}
        try:
            row.status = 'extracting'
            self.db.commit()

            s3 = S3FileUploadService()
            file_bytes = s3.download_file(row.s3_key) if row.s3_key else None
            if not file_bytes:
                raise ValueError('Could not read the uploaded file from storage')

            parsed = P.parse_setup_workbook(file_bytes)
            preview = self._build_preview(parsed, organization_id)

            row.preview_payload = preview
            row.summary = preview['summary']
            row.status = 'extracted'
            row.error = None
            self.db.commit()
            self.db.refresh(row)
            return {'success': True, 'data': {**self._row_meta(row), **preview}}
        except Exception as e:
            self.db.rollback()
            self._mark_failed(import_id, organization_id, str(e))
            logger.error(f"setup-import extract error: {e}")
            return {'success': False, 'message': 'Extraction failed', 'errors': [str(e)]}

    def get_preview(self, import_id: int, organization_id: int) -> Dict[str, Any]:
        row = self._get_row(import_id, organization_id)
        if not row:
            return {'success': False, 'message': 'Import not found'}
        return {'success': True, 'data': {**self._row_meta(row), **(row.preview_payload or {})}}

    def _build_preview(self, parsed: Dict[str, Any], organization_id: int) -> Dict[str, Any]:
        """Validate every section and produce an editable preview. Blocking errors gate confirm."""
        # Names available to resolve Members against: imported users (by display_name).
        imported_user_names = {_norm(u['display_name']) for u in parsed['users'] if u['display_name']}
        # Existing active org users (members may reference people already in the org).
        existing_users = self.db.query(UserLocation).filter(
            UserLocation.organization_id == organization_id, UserLocation.is_user == True,  # noqa: E712
            UserLocation.is_active == True, UserLocation.deleted_date.is_(None),  # noqa: E712
        ).all()
        existing_user_names = {
            _norm(n) for u in existing_users
            for n in [u.display_name, u.name_en, u.name_th] if n
        }
        # Display labels of existing org users — offered (alongside imported users) as
        # member-chip options in the preview.
        existing_member_labels = sorted({
            (u.display_name or u.name_en or u.name_th)
            for u in existing_users if (u.display_name or u.name_en or u.name_th)
        })
        known_member_names = imported_user_names | existing_user_names

        # Material names → resolvable (global or this org).
        materials = self.db.query(Material).filter(
            Material.is_active == True,  # noqa: E712
            (Material.is_global == True) | (Material.organization_id == organization_id),  # noqa: E712
        ).all()
        known_material_names = {_norm(n) for m in materials for n in [m.name_en, m.name_th] if n}
        # Material display labels — the multi-select options for the Materials column in the preview.
        material_option_labels = sorted({
            (m.name_en or m.name_th) for m in materials if (m.name_en or m.name_th)
        })

        # Existing destination (hub) names → id — a NEW destination must not collide, but an
        # UPSERT row keeping its own name (same id) is fine.
        existing_dest_id_by_name: Dict[str, int] = {}
        for u in self.db.query(UserLocation).filter(
            UserLocation.organization_id == organization_id, UserLocation.hub_type.isnot(None),
            UserLocation.deleted_date.is_(None),
        ).all():
            if u.display_name:
                existing_dest_id_by_name.setdefault(_norm(u.display_name), int(u.id))

        # Owner is excluded from import (never created/updated). Rows resolving to the owner
        # (by id or by email) are dropped from the preview entirely.
        owner_id, owner_email = self._owner_info(organization_id)

        # Existing record ids per type — an ID in the file that matches one of these is an UPSERT
        # (mode='update'); otherwise mode='new'. Also used to validate ID-uniqueness within a type.
        existing_user_ids = {int(u.id) for u in existing_users}
        # email(lower) → this-org live user id — lets an update be matched by EMAIL when the file's
        # ID column is stale/blank (e.g. a prior import re-created the user under a new id), so a
        # re-import stays idempotent instead of flagging a false "email already exists".
        existing_user_id_by_email = {}
        for eu in existing_users:
            if eu.email:
                existing_user_id_by_email.setdefault(eu.email.strip().lower(), int(eu.id))
        existing_tag_ids = {int(r[0]) for r in self.db.query(UserLocationTag.id).filter(
            UserLocationTag.organization_id == organization_id, UserLocationTag.deleted_date.is_(None)).all()}
        existing_tenant_ids = {int(r[0]) for r in self.db.query(UserTenant.id).filter(
            UserTenant.organization_id == organization_id, UserTenant.deleted_date.is_(None)).all()}
        existing_loc_ids = {int(r[0]) for r in self.db.query(UserLocation.id).filter(
            UserLocation.organization_id == organization_id, UserLocation.is_location == True,  # noqa: E712
            UserLocation.deleted_date.is_(None)).all()}

        def _id_check(rid, seen: set, existing: set, errors: list) -> str:
            """Validate an optional ID: duplicate-within-type is blocking; return 'update'/'new'."""
            if not rid:
                return 'new'
            if rid in seen:
                errors.append('duplicate_id')
            seen.add(rid)
            return 'update' if rid in existing else 'new'

        blocking = 0

        # -- Users --
        users_out, seen_emails, seen_user_ids = [], {}, set()
        for u in parsed['users']:
            email = (u['email'] or '').strip()
            email_lower = email.lower()
            rid = u.get('id')

            # Skip the org owner — never imported (create or update).
            if owner_id is not None and (
                (rid is not None and int(rid) == owner_id)
                or (owner_email and email_lower == owner_email)
                or (email and self._email_owner_id(email) == owner_id)
            ):
                continue

            errors = []
            # Resolve which existing user (this org) this row targets: by explicit ID first, then
            # by email. Either match ⇒ mode='update' and we reuse that id (so confirm updates it).
            by_id = int(rid) if (rid is not None and int(rid) in existing_user_ids) else None
            by_email = existing_user_id_by_email.get(email_lower) if email_lower else None
            target_id = by_id if by_id is not None else by_email
            mode = 'update' if target_id is not None else 'new'
            # duplicate-ID within the file (only meaningful for an explicit ID)
            if rid is not None:
                if int(rid) in seen_user_ids:
                    errors.append('duplicate_id')
                seen_user_ids.add(int(rid))

            if not email:
                errors.append('email_required')
            elif not _EMAIL_RE.match(email):
                errors.append('email_invalid')
            else:
                if email_lower in seen_emails:
                    errors.append('email_duplicate_in_file')
                seen_emails[email_lower] = True
                # An email is a blocking duplicate ONLY when a LIVE user other than the one this
                # row updates owns it. Same owner as target_id (or target_id itself) ⇒ update, fine.
                owner_of_email = self._email_owner_id(email)
                if owner_of_email is not None and owner_of_email != target_id:
                    errors.append('email_exists')  # belongs to a different existing user
            role_key = _role_key(u['role'])
            if not role_key:
                errors.append('role_unknown')
            if not u['display_name']:
                errors.append('display_name_required')
            # Blank password → generate one ONLY for a new user (an update keeps the existing
            # password unless the admin types a new one). Stored in the preview so it's visible/
            # exportable, and confirm reuses this exact value.
            pwd = u.get('password')
            password_generated = (not pwd) and mode == 'new'
            if password_generated:
                pwd = _gen_password()
            blocking += len(errors)
            users_out.append({
                **u, 'id': target_id, 'password': pwd, 'password_generated': password_generated,
                'role_key': role_key, 'mode': mode, 'errors': errors,
            })

        # -- Tags / Tenants (INSERT or UPSERT-by-id; only name required; members are warnings) --
        def _named(section, existing_ids):
            out, seen_ids = [], set()
            for t in parsed[section]:
                errors, warnings = [], []
                mode = _id_check(t.get('id'), seen_ids, existing_ids, errors)
                if not t['name']:
                    errors.append('name_required')
                unmatched = [m for m in t['members'] if _norm(m) not in known_member_names]
                if unmatched:
                    warnings.append('members_unmatched')
                out.append({**t, 'unmatched_members': unmatched, 'mode': mode,
                            'errors': errors, 'warnings': warnings})
                nonlocal_blocking[0] += len(errors)
            return out
        nonlocal_blocking = [0]
        tags_out = _named('tags', existing_tag_ids)
        tenants_out = _named('tenants', existing_tenant_ids)
        blocking += nonlocal_blocking[0]

        # -- Origins (REPLACE) — sibling-name uniqueness + reference checks --
        origins_out = []
        by_parent: Dict[Tuple[str, ...], set] = {}
        seen_origin_ids: set = set()
        origin_tag_names = {_norm(t['name']) for t in parsed['tags'] if t['name']}
        origin_tenant_names = {_norm(t['name']) for t in parsed['tenants'] if t['name']}
        for o in parsed['origins']:
            errors, warnings = [], []
            mode = _id_check(o.get('id'), seen_origin_ids, existing_loc_ids, errors)
            if o['malformed']:
                errors.append('malformed_path')
            # Siblings (same parent path) must have unique names; different parents may repeat.
            pkey = tuple(_norm(x) for x in o['parent_path'])
            sib = by_parent.setdefault(pkey, set())
            nkey = _norm(o['name'])
            if nkey in sib:
                errors.append('duplicate_sibling')
            sib.add(nkey)
            if [m for m in o['members'] if _norm(m) not in known_member_names]:
                warnings.append('members_unmatched')
            if [t for t in o['tags'] if _norm(t) not in origin_tag_names]:
                warnings.append('tags_unmatched')
            if [t for t in o['tenants'] if _norm(t) not in origin_tenant_names]:
                warnings.append('tenants_unmatched')
            if [m for m in o['materials'] if _norm(m) not in known_material_names]:
                warnings.append('materials_unmatched')
            blocking += len(errors)
            origins_out.append({**o, 'mode': mode, 'errors': errors, 'warnings': warnings})

        # -- Destinations (REPLACE or UPSERT-by-id) — unique name (vs existing + in-file) --
        dests_out, seen_dest, seen_dest_ids = [], [], set()
        for d in parsed['destinations']:
            errors, warnings = [], []
            rid = d.get('id')
            mode = _id_check(rid, seen_dest_ids, existing_loc_ids, errors)
            if not d['name']:
                errors.append('name_required')
            nkey = _norm(d['name'])
            if nkey in seen_dest:
                errors.append('duplicate_in_file')
            seen_dest.append(nkey)
            # A name that already exists on a DIFFERENT destination collides; keeping your own
            # name (same id) is fine.
            if existing_dest_id_by_name.get(nkey) not in (None, rid):
                errors.append('name_exists')
            if [m for m in d['members'] if _norm(m) not in known_member_names]:
                warnings.append('members_unmatched')
            if [m for m in d['materials'] if _norm(m) not in known_material_names]:
                warnings.append('materials_unmatched')
            blocking += len(errors)
            dests_out.append({**d, 'hub_type_key': _hub_type_key(d['business_type']),
                              'mode': mode, 'errors': errors, 'warnings': warnings})

        summary = {
            'users': len(users_out), 'tags': len(tags_out), 'tenants': len(tenants_out),
            'origins': len(origins_out), 'destinations': len(dests_out),
            'blocking_errors': blocking,
        }
        return {
            'users': users_out, 'tags': tags_out, 'tenants': tenants_out,
            'origins': origins_out, 'destinations': dests_out,
            'existing_members': existing_member_labels,  # + imported users → member-chip options
            'material_options': material_option_labels,  # options for the Materials multi-select
            'summary': summary, 'can_confirm': blocking == 0,
        }

    # ── Export current setup (re-importable, with real IDs) ──────────────────────
    def export_setup(self, organization_id: int) -> Dict[str, Any]:
        """
        Export the org's CURRENT setup as an .xlsx matching Organization Setup.xlsx, plus a
        trailing **ID** column per tab (the real record id). Re-importing a file with IDs upserts
        (reuse/update) those records instead of creating duplicates. References (Members/Tag/
        Tenant/Materials) are emitted as NAMES, same as the import format.
        """
        import io
        import base64
        from openpyxl import Workbook

        # Owner is excluded from the export — every other user is exported, never the owner.
        owner_id, _owner_email = self._owner_info(organization_id)

        setup = self.db.query(OrganizationSetup).filter(
            OrganizationSetup.organization_id == organization_id,
            OrganizationSetup.is_active == True,  # noqa: E712
        ).order_by(OrganizationSetup.created_date.desc()).first()
        if not setup:
            setup = self.db.query(OrganizationSetup).filter(
                OrganizationSetup.organization_id == organization_id
            ).order_by(OrganizationSetup.created_date.desc()).first()
        root_nodes = (setup.root_nodes if setup and isinstance(setup.root_nodes, list) else []) or []
        hub_node = (setup.hub_node if setup and isinstance(setup.hub_node, dict) else {}) or {}

        locs = self.db.query(UserLocation).filter(
            UserLocation.organization_id == organization_id,
            UserLocation.is_active == True, UserLocation.deleted_date.is_(None),  # noqa: E712
        ).all()
        loc_by_id = {int(l.id): l for l in locs}
        user_name_by_id = {int(l.id): (l.display_name or l.name_en or l.name_th or str(l.id))
                           for l in locs if l.is_user}

        tag_rows = self.db.query(UserLocationTag).filter(
            UserLocationTag.organization_id == organization_id, UserLocationTag.deleted_date.is_(None),
            UserLocationTag.is_active == True,  # noqa: E712
        ).all()
        tenant_rows = self.db.query(UserTenant).filter(
            UserTenant.organization_id == organization_id, UserTenant.deleted_date.is_(None),
            UserTenant.is_active == True,  # noqa: E712
        ).all()
        tag_name_by_id = {int(t.id): t.name for t in tag_rows}
        tenant_name_by_id = {int(t.id): t.name for t in tenant_rows}

        mat_name_by_id = {}
        for m in self.db.query(Material).filter(
            Material.is_active == True,  # noqa: E712
            (Material.is_global == True) | (Material.organization_id == organization_id),  # noqa: E712
        ).all():
            mat_name_by_id[int(m.id)] = m.name_en or m.name_th

        # user_location member id-list → display names.
        def member_names(members):
            out = []
            for m in members or []:
                uid = m.get('user_id') if isinstance(m, dict) else m
                try:
                    uid = int(uid)
                except (TypeError, ValueError):
                    continue
                if uid in user_name_by_id:
                    out.append(user_name_by_id[uid])
            return ', '.join(out)

        def id_names(ids, name_map):
            out = []
            for i in ids or []:
                try:
                    i = int(i)
                except (TypeError, ValueError):
                    continue
                if i in name_map:
                    out.append(name_map[i])
            return ', '.join(out)

        def fmt_date(d):
            return d.date().isoformat() if d else ''

        # role_id → key → label (for the Users tab).
        role_key_by_id = {}
        role_ids = {int(l.organization_role_id) for l in locs if l.is_user and l.organization_role_id}
        if role_ids:
            from ....models.subscriptions.subscription_models import OrganizationRole
            for r in self.db.query(OrganizationRole).filter(OrganizationRole.id.in_(list(role_ids))).all():
                role_key_by_id[int(r.id)] = r.key

        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws.title = 'Users'
        ws.append(['Display Name', 'Email', 'Password', 'Role', 'First Name', 'Last Name', 'QR Name', 'ID'])
        for l in sorted((x for x in locs if x.is_user), key=lambda x: int(x.id)):
            if owner_id is not None and int(l.id) == owner_id:
                continue  # never export the owner
            role_label = _ROLE_LABEL.get(role_key_by_id.get(int(l.organization_role_id) if l.organization_role_id else None), '')
            ws.append([l.display_name, l.email, '', role_label, l.first_name, l.last_name, l.qr_name, int(l.id)])
        # Role column (D) is a dropdown in Excel — the 4 valid roles. Covers extra rows the admin
        # may add. The importer still normalises the text, so a typed value also works.
        role_dv = DataValidation(
            type='list', formula1='"' + ','.join(_ROLE_LABEL.values()) + '"', allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(role_dv)
        role_dv.add('D2:D1000')

        for sheet_name, name_col, rows in (
            ('Tags', 'Tag', tag_rows), ('Tenants', 'Tenant', tenant_rows),
        ):
            s = wb.create_sheet(sheet_name)
            s.append([name_col, 'Description', 'Start', 'End', 'Members', 'ID'])
            for t in sorted(rows, key=lambda x: int(x.id)):
                s.append([t.name, t.note, fmt_date(t.start_date), fmt_date(t.end_date),
                          id_names(t.members, user_name_by_id), int(t.id)])

        # Origins — walk root_nodes, one row per node, level path from ancestor names.
        s = wb.create_sheet('Origins')
        s.append(['Level 1\n(Branch)', 'Level 2\n(Building)', 'Level 3\n(Floor)', 'Level 4\n(Room)',
                  'Is Destination', 'Tag\n(Event)', 'Tenant\n(Company)', 'Members', 'Address', 'Materials', 'ID'])

        def _walk_origins(nodes, name_path):
            for node in nodes if isinstance(nodes, list) else []:
                nid = node.get('nodeId')
                try:
                    nid = int(nid)
                except (TypeError, ValueError):
                    continue
                loc = loc_by_id.get(nid)
                if not loc:
                    continue
                nm = loc.display_name or loc.name_en or loc.name_th or str(nid)
                path = name_path + [nm]
                levels = (path + ['', '', '', ''])[:4]
                s.append([
                    *levels,
                    bool(node.get('is_destination')),
                    id_names(loc.tags, tag_name_by_id), id_names(loc.tenants, tenant_name_by_id),
                    member_names(loc.members), loc.address or '',
                    id_names(loc.materials, mat_name_by_id), nid,
                ])
                _walk_origins(node.get('children') or [], path)

        _walk_origins(root_nodes, [])

        # Destinations — hub_node children.
        s = wb.create_sheet('Destination')
        s.append(['Destinations', 'Members', 'Address', 'Business Type', 'Materials', 'ID'])
        for node in (hub_node.get('children') or []):
            try:
                nid = int(node.get('nodeId'))
            except (TypeError, ValueError):
                continue
            loc = loc_by_id.get(nid)
            if not loc:
                continue
            s.append([
                loc.display_name or loc.name_en or loc.name_th or str(nid),
                member_names(loc.members), loc.address or '', loc.business_type or '',
                id_names(loc.materials, mat_name_by_id), nid,
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return {
            'success': True,
            'data': {
                'filename': f'organization_setup_{organization_id}.xlsx',
                'content_base64': base64.b64encode(buf.getvalue()).decode('ascii'),
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            },
        }

    # ── 3. Confirm → create everything ───────────────────────────────────────────
    def confirm(self, import_id: int, organization_id: int, admin_id: Optional[int],
                edited: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        row = self._get_row(import_id, organization_id)
        if not row:
            return {'success': False, 'message': 'Import not found'}
        if row.status == 'confirmed':
            return {'success': False, 'message': 'This import has already been confirmed'}

        preview = edited or row.preview_payload or {}
        # Re-validate emails on the (possibly edited) payload — the admin may have fixed dups.
        revalidation = self._revalidate_users(preview.get('users', []), organization_id)
        if revalidation:
            return {'success': False, 'message': 'Validation failed', 'errors': revalidation}

        created = {'users': [], 'tags': [], 'tenants': [], 'locations': []}
        try:
            row.status = 'confirming'
            self.db.commit()

            from ...cores.users.user_service import UserService
            from ...cores.users.location_tag_service import LocationTagService
            from ...cores.users.tenant_service import TenantService
            usvc, tagsvc, tensvc = UserService(self.db), LocationTagService(self.db), TenantService(self.db)

            # Existing records by id — an ID in the file UPSERTS (updates/reuses) that record
            # instead of creating a new one. Updated records are NOT added to `created` (revert
            # removes only what the import created, never touches reused rows).
            existing_user_by_id = {int(l.id): l for l in self.db.query(UserLocation).filter(
                UserLocation.organization_id == organization_id, UserLocation.is_user == True,  # noqa: E712
                UserLocation.is_active == True, UserLocation.deleted_date.is_(None)).all()}  # noqa: E712
            existing_tag_by_id = {int(t.id): t for t in self.db.query(UserLocationTag).filter(
                UserLocationTag.organization_id == organization_id, UserLocationTag.deleted_date.is_(None)).all()}
            existing_tenant_by_id = {int(t.id): t for t in self.db.query(UserTenant).filter(
                UserTenant.organization_id == organization_id, UserTenant.deleted_date.is_(None)).all()}
            existing_loc_by_id = {int(l.id): l for l in self.db.query(UserLocation).filter(
                UserLocation.organization_id == organization_id, UserLocation.is_location == True,  # noqa: E712
                UserLocation.is_active == True, UserLocation.deleted_date.is_(None)).all()}  # noqa: E712

            # Owner is excluded from import (defensive — the preview already drops owner rows).
            owner_id, owner_email = self._owner_info(organization_id)

            # 1) Users (UPSERT by id). display_name(norm) → id (+ role for location membership).
            user_by_name: Dict[str, int] = {}
            user_role_by_name: Dict[str, str] = {}
            for u in preview.get('users', []):
                _uemail = (u.get('email') or '').strip().lower()
                if owner_id is not None and (
                    (u.get('id') is not None and int(u['id']) == owner_id)
                    or (owner_email and _uemail == owner_email)
                ):
                    continue  # never create/update the owner
                role_key = u.get('role_key') or _role_key(u.get('role')) or 'viewer'
                existing = existing_user_by_id.get(int(u['id'])) if u.get('id') else None
                if existing:
                    # Update in place; not tracked as created.
                    if u.get('display_name'):
                        existing.display_name = u['display_name']
                        existing.name_en = u['display_name']
                    email = (u.get('email') or '').strip()
                    if email:
                        existing.email = email
                    existing.first_name = u.get('first_name')
                    existing.last_name = u.get('last_name')
                    existing.qr_name = u.get('qr_name')
                    rid = usvc.crud._get_organization_role_id(role_key, organization_id)
                    if rid:
                        existing.organization_role_id = rid
                    if u.get('password'):  # only change password when a value is provided
                        existing.password = usvc.crud._hash_password(u['password'])
                    self.db.flush()
                    uid = int(existing.id)
                else:
                    # created_by_id=None: a back-office admin import into a target org the admin
                    # isn't a member of — passing them as creator would trip create_user's
                    # org-inheritance validation and override the explicit organization_id.
                    res = usvc.create_user({
                        'display_name': u['display_name'], 'email': (u['email'] or '').strip(),
                        'password': u.get('password') or None, 'role': role_key,
                        'first_name': u.get('first_name'), 'last_name': u.get('last_name'),
                        'qr_name': u.get('qr_name'), 'organization_id': organization_id,
                        'is_user': True,
                    }, created_by_id=None, auto_generate_credentials=True)
                    if not res.get('success'):
                        raise RuntimeError(res.get('message') or 'User creation failed')
                    uid = int(res['user']['id'])
                    created['users'].append(uid)
                if u.get('display_name'):
                    user_by_name[_norm(u['display_name'])] = uid
                    user_role_by_name[_norm(u['display_name'])] = role_key

            # Existing org users can also be referenced as members.
            for eu in existing_user_by_id.values():
                for n in [eu.display_name, eu.name_en, eu.name_th]:
                    if n:
                        user_by_name.setdefault(_norm(n), int(eu.id))

            # Location members carry {user_id, role} (role = the user's imported role, default
            # viewer for existing-org members) so the business app renders them without crashing.
            def member_objs(names): return [
                {'user_id': user_by_name[_norm(n)], 'role': user_role_by_name.get(_norm(n), 'viewer')}
                for n in names if _norm(n) in user_by_name]
            def member_ids(names): return [user_by_name[_norm(n)] for n in names if _norm(n) in user_by_name]

            # 2) Tags (UPSERT by id). name(norm) → id.
            tag_by_name: Dict[str, int] = {}
            for t in preview.get('tags', []):
                existing = existing_tag_by_id.get(int(t['id'])) if t.get('id') else None
                if existing:
                    existing.name = t['name']
                    existing.note = t.get('description')
                    existing.start_date = _to_dt(t.get('start_date'))
                    existing.end_date = _to_dt(t.get('end_date'))
                    existing.members = member_ids(t.get('members', []))
                    self.db.flush()
                    tid = int(existing.id)
                else:
                    res = tagsvc.create_tag(organization_id, {
                        'name': t['name'], 'note': t.get('description'),
                        'start_date': t.get('start_date'), 'end_date': t.get('end_date'),
                        'members': member_ids(t.get('members', [])),
                    }, created_by_id=admin_id)
                    tid = self._extract_id(res)
                    created['tags'].append(tid)
                tag_by_name[_norm(t['name'])] = tid

            # 3) Tenants (UPSERT by id). name(norm) → id.
            tenant_by_name: Dict[str, int] = {}
            for t in preview.get('tenants', []):
                existing = existing_tenant_by_id.get(int(t['id'])) if t.get('id') else None
                if existing:
                    existing.name = t['name']
                    existing.note = t.get('description')
                    existing.start_date = _to_dt(t.get('start_date'))
                    existing.end_date = _to_dt(t.get('end_date'))
                    existing.members = member_ids(t.get('members', []))
                    self.db.flush()
                    tid = int(existing.id)
                else:
                    res = tensvc.create_tenant(organization_id, {
                        'name': t['name'], 'note': t.get('description'),
                        'start_date': t.get('start_date'), 'end_date': t.get('end_date'),
                        'members': member_ids(t.get('members', [])),
                    }, created_by_id=admin_id)
                    tid = self._extract_id(res)
                    created['tenants'].append(tid)
                tenant_by_name[_norm(t['name'])] = tid

            # Material name → id.
            mat_by_name: Dict[str, int] = {}
            for m in self.db.query(Material).filter(
                Material.is_active == True,  # noqa: E712
                (Material.is_global == True) | (Material.organization_id == organization_id),  # noqa: E712
            ).all():
                for n in [m.name_en, m.name_th]:
                    if n:
                        mat_by_name.setdefault(_norm(n), int(m.id))
            def material_ids(names): return [mat_by_name[_norm(n)] for n in names if _norm(n) in mat_by_name]

            def _tag_ids(names): return [tag_by_name[_norm(t)] for t in names if _norm(t) in tag_by_name]
            def _tenant_ids(names): return [tenant_by_name[_norm(t)] for t in names if _norm(t) in tenant_by_name]

            # 4) Origins (REPLACE tree; UPSERT node rows by id). Parents-first for tree wiring.
            origins = sorted(preview.get('origins', []), key=lambda o: o.get('depth', 1))
            node_by_path: Dict[Tuple[str, ...], Dict[str, Any]] = {}
            root_nodes: List[Dict[str, Any]] = []
            for o in origins:
                existing = existing_loc_by_id.get(int(o['id'])) if o.get('id') else None
                if existing:
                    existing.display_name = o['name']
                    existing.name_en = o['name']
                    existing.name_th = o['name']
                    existing.type = o['type']
                    existing.hub_type = None  # an origin is not a hub
                    existing.address = o.get('address') or None
                    existing.members = member_objs(o.get('members', []))
                    existing.tags = _tag_ids(o.get('tags', []))
                    existing.tenants = _tenant_ids(o.get('tenants', []))
                    existing.materials = material_ids(o.get('materials', []))
                    self.db.flush()
                    nid = int(existing.id)
                else:
                    loc = UserLocation(
                        organization_id=organization_id, is_location=True, is_user=False,
                        display_name=o['name'], name_en=o['name'], name_th=o['name'],
                        type=o['type'], address=o.get('address') or None,
                        members=member_objs(o.get('members', [])),
                        tags=_tag_ids(o.get('tags', [])), tenants=_tenant_ids(o.get('tenants', [])),
                        materials=material_ids(o.get('materials', [])),
                        platform='GEPP_BUSINESS_WEB',
                    )
                    self.db.add(loc)
                    self.db.flush()
                    nid = int(loc.id)
                    created['locations'].append(nid)
                node = {'nodeId': nid, 'children': []}
                if o.get('is_destination'):
                    node['is_destination'] = True
                node_by_path[tuple(_norm(x) for x in o['path'])] = node
                parent = node_by_path.get(tuple(_norm(x) for x in o['parent_path'])) if o.get('parent_path') else None
                (parent['children'] if parent else root_nodes).append(node)

            # 5) Destinations (REPLACE; UPSERT by id) → hub_node.children.
            hub_children: List[Dict[str, Any]] = []
            for d in preview.get('destinations', []):
                htype = d.get('hub_type_key') or _hub_type_key(d.get('business_type', ''))
                existing = existing_loc_by_id.get(int(d['id'])) if d.get('id') else None
                if existing:
                    existing.display_name = d['name']
                    existing.name_en = d['name']
                    existing.name_th = d['name']
                    existing.type = 'hub'
                    existing.hub_type = htype
                    existing.business_type = d.get('business_type') or None
                    existing.address = d.get('address') or None
                    existing.members = member_objs(d.get('members', []))
                    existing.materials = material_ids(d.get('materials', []))
                    self.db.flush()
                    nid = int(existing.id)
                else:
                    loc = UserLocation(
                        organization_id=organization_id, is_location=True, is_user=False,
                        display_name=d['name'], name_en=d['name'], name_th=d['name'],
                        type='hub', hub_type=htype,
                        business_type=d.get('business_type') or None, address=d.get('address') or None,
                        members=member_objs(d.get('members', [])),
                        materials=material_ids(d.get('materials', [])),
                        platform='GEPP_BUSINESS_WEB',
                    )
                    self.db.add(loc)
                    self.db.flush()
                    nid = int(loc.id)
                    created['locations'].append(nid)
                hub_children.append({'nodeId': nid, 'children': []})

            # 6) New setup version (REPLACE) — carry over level names; trigger deactivates prior.
            prev = self.db.query(OrganizationSetup).filter(
                OrganizationSetup.organization_id == organization_id,
            ).order_by(OrganizationSetup.created_date.desc()).first()
            hub_node = dict(prev.hub_node) if (prev and isinstance(prev.hub_node, dict)) else {}
            hub_node['children'] = hub_children
            new_setup = OrganizationSetup(
                organization_id=organization_id, version=self._next_version(prev), is_active=True,
                root_nodes=root_nodes, hub_node=hub_node,
                branch_level_name=getattr(prev, 'branch_level_name', None),
                building_level_name=getattr(prev, 'building_level_name', None),
                floor_level_name=getattr(prev, 'floor_level_name', None),
                room_level_name=getattr(prev, 'room_level_name', None),
            )
            self.db.add(new_setup)
            self.db.flush()

            row.created_user_ids = created['users']
            row.created_tag_ids = created['tags']
            row.created_tenant_ids = created['tenants']
            row.created_location_ids = created['locations']
            row.created_setup_version_id = int(new_setup.id)
            row.status = 'confirmed'
            row.confirmed_date = datetime.now()
            row.summary = {**(row.summary or {}), 'created': {
                'users': len(created['users']), 'tags': len(created['tags']),
                'tenants': len(created['tenants']), 'locations': len(created['locations']),
            }}
            self.db.commit()
            self.db.refresh(row)
            return {'success': True, 'data': self._row_meta(row)}
        except Exception as e:
            self.db.rollback()
            logger.error(f"setup-import confirm error: {e} — compensating")
            self._soft_delete(created)  # remove anything already created (individual commits)
            self._mark_failed(import_id, organization_id, str(e))
            return {'success': False, 'message': 'Import failed and was rolled back', 'errors': [str(e)]}

    # ── 4. History + revert ──────────────────────────────────────────────────────
    def list_history(self, organization_id: int) -> Dict[str, Any]:
        rows = self.db.query(OrganizationSetupImport).filter(
            OrganizationSetupImport.organization_id == organization_id,
            OrganizationSetupImport.deleted_date.is_(None),
        ).order_by(OrganizationSetupImport.created_date.desc()).all()
        return {'success': True, 'data': [self._row_meta(r) for r in rows]}

    def revert(self, import_id: int, organization_id: int) -> Dict[str, Any]:
        row = self._get_row(import_id, organization_id)
        if not row:
            return {'success': False, 'message': 'Import not found'}
        if row.status != 'confirmed':
            return {'success': False, 'message': 'Only a confirmed import can be reverted'}
        try:
            created_locs = set(int(x) for x in (row.created_location_ids or []))
            # Soft-delete created users/tags/tenants/locations.
            self._soft_delete({
                'users': row.created_user_ids or [], 'tags': row.created_tag_ids or [],
                'tenants': row.created_tenant_ids or [], 'locations': list(created_locs),
            })
            # Strip the imported nodeIds from the active tree → new version (leaves the
            # pre-import orphans in the recycle bin, untouched).
            active = self.db.query(OrganizationSetup).filter(
                OrganizationSetup.organization_id == organization_id,
                OrganizationSetup.is_active == True,  # noqa: E712
            ).order_by(OrganizationSetup.created_date.desc()).first()
            if active:
                new_root = self._strip_nodes(active.root_nodes or [], created_locs)
                hub = dict(active.hub_node) if isinstance(active.hub_node, dict) else {}
                hub['children'] = self._strip_nodes(hub.get('children') or [], created_locs)
                new_setup = OrganizationSetup(
                    organization_id=organization_id, version=self._next_version(active), is_active=True,
                    root_nodes=new_root, hub_node=hub,
                    branch_level_name=active.branch_level_name, building_level_name=active.building_level_name,
                    floor_level_name=active.floor_level_name, room_level_name=active.room_level_name,
                )
                self.db.add(new_setup)
                self.db.flush()
            row.status = 'reverted'
            row.reverted_date = datetime.now()
            self.db.commit()
            self.db.refresh(row)
            return {'success': True, 'data': self._row_meta(row)}
        except Exception as e:
            self.db.rollback()
            logger.error(f"setup-import revert error: {e}")
            return {'success': False, 'message': 'Revert failed', 'errors': [str(e)]}

    # ── internals ────────────────────────────────────────────────────────────────
    def _get_row(self, import_id: int, organization_id: int) -> Optional[OrganizationSetupImport]:
        return self.db.query(OrganizationSetupImport).filter(
            OrganizationSetupImport.id == import_id,
            OrganizationSetupImport.organization_id == organization_id,
            OrganizationSetupImport.deleted_date.is_(None),
        ).first()

    def _mark_failed(self, import_id: int, organization_id: int, error: str) -> None:
        try:
            r = self._get_row(import_id, organization_id)
            if r:
                r.status = 'failed'
                r.error = (error or '')[:2000]
                self.db.commit()
        except Exception:
            self.db.rollback()

    def _email_exists(self, email: str, exclude_id: Optional[int] = None) -> bool:
        """
        Global email uniqueness among LIVE users. A soft-deleted user (deleted_date set OR
        is_active=False) frees its email — it can be imported again. exclude_id lets an upsert
        row keep its own current email.
        """
        from sqlalchemy import func
        q = self.db.query(UserLocation.id).filter(
            func.lower(UserLocation.email) == email.strip().lower(),
            UserLocation.is_user == True,  # noqa: E712
            UserLocation.is_active == True,  # noqa: E712 — deleted-by-flag users don't block
            UserLocation.deleted_date.is_(None),
        )
        if exclude_id is not None:
            q = q.filter(UserLocation.id != int(exclude_id))
        return q.first() is not None

    def _email_owner_id(self, email: str) -> Optional[int]:
        """Live global owner (user_location id) of an email, or None. Global uniqueness ⇒ ≤1."""
        if not email:
            return None
        from sqlalchemy import func
        r = self.db.query(UserLocation.id).filter(
            func.lower(UserLocation.email) == email.strip().lower(),
            UserLocation.is_user == True,  # noqa: E712
            UserLocation.is_active == True,  # noqa: E712
            UserLocation.deleted_date.is_(None),
        ).order_by(UserLocation.id.asc()).first()
        return int(r[0]) if r else None

    def _owner_info(self, organization_id: int) -> Tuple[Optional[int], Optional[str]]:
        """(owner_user_id, owner_email_lower) for an org. The owner is excluded from
        import AND export — you can import/export every other user, never the owner."""
        row = self.db.query(Organization.owner_id).filter(Organization.id == organization_id).first()
        owner_id = int(row[0]) if row and row[0] is not None else None
        owner_email = None
        if owner_id is not None:
            er = self.db.query(UserLocation.email).filter(UserLocation.id == owner_id).first()
            owner_email = (er[0].strip().lower() if er and er[0] else None)
        return owner_id, owner_email

    def _revalidate_users(self, users: List[Dict[str, Any]],
                          organization_id: Optional[int] = None) -> List[str]:
        errors, seen = [], set()
        owner_id, owner_email = self._owner_info(organization_id) if organization_id else (None, None)
        for u in users:
            email = (u.get('email') or '').strip()
            # Owner rows are never imported → skip (defensive; preview already drops them).
            if owner_id is not None and (
                (u.get('id') is not None and int(u['id']) == owner_id)
                or (owner_email and email.lower() == owner_email)
            ):
                continue
            if not email or not _EMAIL_RE.match(email):
                errors.append(f"Row {u.get('row_index')}: invalid email")
                continue
            k = email.lower()
            if k in seen:
                errors.append(f"Row {u.get('row_index')}: duplicate email in file ({email})")
            seen.add(k)
            # An update row reuses its own id → owning that email is fine; only a DIFFERENT
            # live user owning it blocks (exclude_id handles the self-ownership case).
            if self._email_exists(email, exclude_id=u.get('id')):
                errors.append(f"Row {u.get('row_index')}: email already exists ({email})")
            if not (u.get('role_key') or _role_key(u.get('role'))):
                errors.append(f"Row {u.get('row_index')}: unknown role")
        return errors

    @staticmethod
    def _extract_id(res: Any) -> int:
        """Pull the created id from a create_tag/create_tenant result (tolerant of envelope shape)."""
        if isinstance(res, dict):
            for k in ('data', 'tag', 'tenant'):
                v = res.get(k)
                if isinstance(v, dict) and v.get('id') is not None:
                    return int(v['id'])
            if res.get('id') is not None:
                return int(res['id'])
        return int(getattr(res, 'id'))

    def _soft_delete(self, ids: Dict[str, List[int]]) -> None:
        now = datetime.now()
        loc_ids = [int(x) for x in (ids.get('users') or [])] + [int(x) for x in (ids.get('locations') or [])]
        if loc_ids:
            self.db.query(UserLocation).filter(UserLocation.id.in_(loc_ids)).update(
                {UserLocation.is_active: False, UserLocation.deleted_date: now}, synchronize_session=False)
        if ids.get('tags'):
            self.db.query(UserLocationTag).filter(UserLocationTag.id.in_([int(x) for x in ids['tags']])).update(
                {UserLocationTag.is_active: False, UserLocationTag.deleted_date: now}, synchronize_session=False)
        if ids.get('tenants'):
            self.db.query(UserTenant).filter(UserTenant.id.in_([int(x) for x in ids['tenants']])).update(
                {UserTenant.is_active: False, UserTenant.deleted_date: now}, synchronize_session=False)
        self.db.commit()

    @staticmethod
    def _strip_nodes(nodes: List[Dict[str, Any]], remove_ids: set) -> List[Dict[str, Any]]:
        """Drop any node whose nodeId is in remove_ids (its whole subtree goes with it)."""
        out = []
        for n in nodes or []:
            try:
                nid = int(n.get('nodeId'))
            except (TypeError, ValueError):
                nid = None
            if nid in remove_ids:
                continue
            kept = dict(n)
            kept['children'] = SetupImportService._strip_nodes(n.get('children') or [], remove_ids)
            out.append(kept)
        return out

    @staticmethod
    def _next_version(prev) -> str:
        if not prev:
            return '1.0'
        try:
            return str(round(float(prev.version) + 0.1, 1))
        except (ValueError, TypeError):
            return '1.1'

    def _row_meta(self, r: OrganizationSetupImport) -> Dict[str, Any]:
        return {
            'id': r.id, 'organization_id': r.organization_id, 'status': r.status,
            'original_filename': r.original_filename, 'file_size': r.file_size,
            'summary': r.summary, 'error': r.error,
            'created_date': r.created_date.isoformat() if r.created_date else None,
            'confirmed_date': r.confirmed_date.isoformat() if r.confirmed_date else None,
            'reverted_date': r.reverted_date.isoformat() if r.reverted_date else None,
        }
