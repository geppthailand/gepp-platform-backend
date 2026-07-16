"""
Parse the "Organization Setup" workbook (5 tabs) into structured, ready-to-validate data.

Tabs (tolerant name match): Users, Tags, Tenants, Origins, Destination(s).
Dependency-light: only openpyxl + stdlib. Reference fields (Members / Tag / Tenant / Materials)
are returned as **name lists** — resolving names → ids and validating happens in the service layer.

Origins are one-node-per-row keyed by the Level 1-4 path (contiguous non-blank prefix); depth →
type (1 branch, 2 building, 3 floor, 4 room). The parser records each node's `path` + `parent_path`
so the service can build the root_nodes tree and enforce sibling-name uniqueness.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional


# ── cell helpers ──────────────────────────────────────────────────────────────
def _clean(value: Any) -> str:
    return '' if value is None else str(value).strip()


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    return s == '' or s == '-' or s == '—'


def _split_names(value: Any) -> List[str]:
    """Comma-separated cell → trimmed, de-duplicated (order-preserving) name list."""
    if _is_blank(value):
        return []
    out: List[str] = []
    seen = set()
    for part in str(value).split(','):
        name = part.strip()
        if name and name.lower() not in seen:
            seen.add(name.lower())
            out.append(name)
    return out


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ('true', '1', 'yes', 'y', 't', 'x', '✓', 'checked')


def _parse_date(value: Any) -> Optional[str]:
    """Return an ISO date string (YYYY-MM-DD) or None. Tolerant: unparseable/invalid → None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace('Z', '')).date().isoformat()
    except Exception:
        return None  # e.g. "2026-06-31" (invalid day) → dropped rather than crashing


def _norm(text: str) -> str:
    return re.sub(r'\s+', '', str(text or '').strip().lower())


# Depth (1-based count of non-blank levels) → location type.
_DEPTH_TYPE = {1: 'branch', 2: 'building', 3: 'floor', 4: 'room'}


def _sheet(wb, *names: str):
    """Find a sheet by exact then normalized name; None if absent."""
    for want in names:
        if want in wb.sheetnames:
            return wb[want]
    wanted = {_norm(n) for n in names}
    for sn in wb.sheetnames:
        if _norm(sn) in wanted:
            return wb[sn]
    return None


def _rows(ws) -> List[List[Any]]:
    if ws is None:
        return []
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _cell(row: List[Any], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def _find_id_col(header_row) -> Optional[int]:
    """Index of the optional 'ID' column (case-insensitive) in the header row, or None.
    Used for upsert: an ID that matches an existing record updates it instead of creating new."""
    for idx, cell in enumerate(header_row or []):
        if _norm(cell) == 'id':
            return idx
    return None


def _parse_id(row: List[Any], id_idx: Optional[int]) -> Optional[int]:
    """Read the ID cell → int, or None when absent / blank / '-' (→ treated as create-new)."""
    if id_idx is None:
        return None
    v = _cell(row, id_idx)
    if _is_blank(v):
        return None
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


# ── section parsers ─────────────────────────────────────────────────────────────
def _parse_users(ws) -> List[Dict[str, Any]]:
    out = []
    rows = _rows(ws)
    id_idx = _find_id_col(rows[0]) if rows else None
    for i, row in enumerate(rows[1:], start=2):  # row 1 = header
        if all(_is_blank(c) for c in row):
            continue
        display = _clean(_cell(row, 0))
        email = _clean(_cell(row, 1))
        if not display and not email:
            continue
        out.append({
            'row_index': i,
            'id': _parse_id(row, id_idx),
            'display_name': display,
            'email': email,
            'password': _clean(_cell(row, 2)) or None,
            'role': _clean(_cell(row, 3)),
            'first_name': _clean(_cell(row, 4)),
            'last_name': _clean(_cell(row, 5)),
            'qr_name': _clean(_cell(row, 6)),
        })
    return out


def _parse_named_group(ws) -> List[Dict[str, Any]]:
    """Shared shape for Tags + Tenants: name, description, start, end, members."""
    out = []
    rows = _rows(ws)
    id_idx = _find_id_col(rows[0]) if rows else None
    for i, row in enumerate(rows[1:], start=2):
        if all(_is_blank(c) for c in row):
            continue
        name = _clean(_cell(row, 0))
        if not name:
            continue
        out.append({
            'row_index': i,
            'id': _parse_id(row, id_idx),
            'name': name,
            'description': _clean(_cell(row, 1)),
            'start_date': _parse_date(_cell(row, 2)),
            'end_date': _parse_date(_cell(row, 3)),
            'members': _split_names(_cell(row, 4)),
        })
    return out


def _parse_origins(ws) -> List[Dict[str, Any]]:
    """
    Columns: L1(Branch), L2(Building), L3(Floor), L4(Room), Is Destination, Tag, Tenant,
    Members, Address, Materials. One node per row keyed by the contiguous non-blank level path.
    `malformed=True` when a level is blank then a deeper level is filled (gap in the path).
    """
    out = []
    rows = _rows(ws)
    id_idx = _find_id_col(rows[0]) if rows else None
    for i, row in enumerate(rows[1:], start=2):
        levels_raw = [_cell(row, c) for c in range(4)]
        if all(_is_blank(c) for c in row):
            continue
        # Contiguous non-blank prefix defines the node + its depth.
        path: List[str] = []
        malformed = False
        broke = False
        for c in levels_raw:
            if _is_blank(c):
                broke = True
                continue
            if broke:
                malformed = True  # value after a gap
                break
            path.append(_clean(c))
        if not path:
            continue
        depth = len(path)
        out.append({
            'row_index': i,
            'id': _parse_id(row, id_idx),
            'levels': [_clean(c) for c in levels_raw],
            'path': path,
            'parent_path': path[:-1],
            'depth': depth,
            'type': _DEPTH_TYPE.get(depth, 'room'),
            'name': path[-1],
            'malformed': malformed,
            'is_destination': _parse_bool(_cell(row, 4)),
            'tags': _split_names(_cell(row, 5)),
            'tenants': _split_names(_cell(row, 6)),
            'members': _split_names(_cell(row, 7)),
            'address': _clean(_cell(row, 8)),
            'materials': _split_names(_cell(row, 9)),
        })
    return out


def _parse_destinations(ws) -> List[Dict[str, Any]]:
    """Columns: Destinations(name), Members, Address, Business Type, Materials."""
    out = []
    rows = _rows(ws)
    id_idx = _find_id_col(rows[0]) if rows else None
    for i, row in enumerate(rows[1:], start=2):
        if all(_is_blank(c) for c in row):
            continue
        name = _clean(_cell(row, 0))
        if not name:
            continue
        out.append({
            'row_index': i,
            'id': _parse_id(row, id_idx),
            'name': name,
            'members': _split_names(_cell(row, 1)),
            'address': _clean(_cell(row, 2)),
            'business_type': _clean(_cell(row, 3)),
            'materials': _split_names(_cell(row, 4)),
        })
    return out


def parse_setup_workbook(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse the uploaded .xlsx into structured sections. Raises ValueError if it can't be read.
    Returns { users, tags, tenants, origins, destinations } — all reference fields as name lists.
    """
    import openpyxl  # local import keeps module load cheap

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f'Could not read the workbook: {e}')

    try:
        return {
            'users': _parse_users(_sheet(wb, 'Users')),
            'tags': _parse_named_group(_sheet(wb, 'Tags')),
            'tenants': _parse_named_group(_sheet(wb, 'Tenants')),
            'origins': _parse_origins(_sheet(wb, 'Origins')),
            'destinations': _parse_destinations(_sheet(wb, 'Destination', 'Destinations')),
        }
    finally:
        wb.close()
