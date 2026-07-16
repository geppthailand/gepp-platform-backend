"""
Pure-python helpers for the Excel transaction import: xlsx parsing + name matching.

Kept dependency-light on purpose — only `openpyxl` (for parsing) plus the standard
library. Name matching is char-trigram **cosine similarity** (as specified), computed
by hand rather than via rapidfuzz/sklearn so nothing heavy is added to the Lambda layer.
Matching always picks the argmax candidate — there is **no similarity threshold**.
"""

from __future__ import annotations

import io
import math
import re
from collections import Counter
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple


# ── Blank / sentinel handling ────────────────────────────────────────────────
def is_blank(value: Any) -> bool:
    """A cell counts as blank (→ 'stop here' / 'no value') when empty or a dash."""
    if value is None:
        return True
    s = str(value).strip()
    return s == '' or s == '-' or s == '—' or s == 'N/A' or s.lower() == 'na'


def _clean(value: Any) -> str:
    return '' if value is None else str(value).strip()


# ── Normalisation + char-trigram cosine ──────────────────────────────────────
def _normalize(text: str) -> str:
    """Lowercase + collapse all whitespace (incl. the header newlines) to nothing."""
    return re.sub(r'\s+', '', str(text or '').strip().lower())


def _trigrams(text: str) -> Counter:
    """Padded character trigrams of the normalised text."""
    n = _normalize(text)
    if not n:
        return Counter()
    padded = f'  {n}  '  # pad so short strings still share n-grams
    return Counter(padded[i:i + 3] for i in range(len(padded) - 2))


def cosine_similarity(a: str, b: str) -> float:
    """Cosine similarity of two strings over their char-trigram count vectors (0..1)."""
    va, vb = _trigrams(a), _trigrams(b)
    if not va or not vb:
        return 0.0
    common = set(va) & set(vb)
    dot = sum(va[t] * vb[t] for t in common)
    if dot == 0:
        return 0.0
    mag_a = math.sqrt(sum(v * v for v in va.values()))
    mag_b = math.sqrt(sum(v * v for v in vb.values()))
    return dot / (mag_a * mag_b) if mag_a and mag_b else 0.0


def best_candidate(
    query: str,
    candidates: List[Tuple[Any, List[str]]],
    prefer_flags: Optional[Dict[Any, bool]] = None,
) -> Tuple[Optional[Any], float]:
    """
    Return (key, score) of the candidate whose best name is most similar to `query`.

    candidates: list of (key, [name variants]) — the max similarity across a candidate's
    name variants (e.g. name_th + name_en) is that candidate's score.
    prefer_flags: optional {key: bool}; among equal scores, a True-flagged key wins
    (used to bias waste-type matching toward kilogram-unit materials).

    No threshold: the argmax is always returned (or (None, 0.0) if there are no candidates
    or the query is blank).
    """
    if is_blank(query) or not candidates:
        return None, 0.0
    prefer_flags = prefer_flags or {}
    best_key: Optional[Any] = None
    best_score = -1.0
    best_pref = False
    for key, names in candidates:
        score = max((cosine_similarity(query, nm) for nm in names if nm), default=0.0)
        pref = bool(prefer_flags.get(key, False))
        # Rank by score, then by preferred flag (kg materials) as the tie-breaker.
        if score > best_score or (score == best_score and pref and not best_pref):
            best_key, best_score, best_pref = key, score, pref
    return best_key, max(best_score, 0.0)


# ── Date parsing ──────────────────────────────────────────────────────────────
_DATE_FORMATS = (
    '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M',
    '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
    '%m/%d/%Y %H:%M:%S', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d',
)


def parse_datetime(value: Any) -> Optional[datetime]:
    """
    Parse a cell into a naive datetime (wall-clock). openpyxl already yields datetime for
    date-typed cells; strings are parsed against a set of common formats. A date with no
    time component resolves to 00:00 (the caller treats that as the frontend-tz midnight).
    Returns None if unparseable.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    # openpyxl may hand back a date (not datetime)
    try:
        from datetime import date as _date
        if isinstance(value, _date):
            return datetime(value.year, value.month, value.day)
    except Exception:
        pass
    s = str(value).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Last resort: ISO parser
    try:
        return datetime.fromisoformat(s.replace('Z', ''))
    except Exception:
        return None


def parse_weight(value: Any) -> Optional[float]:
    """Parse a weight cell (handles thousands separators); None if blank/invalid."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(',', '')
    if s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ── xlsx parsing (focus: the "Waste Data" sheet) ──────────────────────────────
# Maps a normalised header (whitespace-stripped, lowercased) to a canonical field.
def _canonical_header(raw: str) -> Optional[str]:
    n = _normalize(raw)
    if n.startswith('date'):
        return 'date'
    if n.startswith('level1'):
        return 'level1'
    if n.startswith('level2'):
        return 'level2'
    if n.startswith('level3'):
        return 'level3'
    if n.startswith('level4'):
        return 'level4'
    if n.startswith('tag'):
        return 'tag'
    if n.startswith('tenant'):
        return 'tenant'
    if n.startswith('wastetype'):
        return 'waste_type'
    if n.startswith('weight'):
        return 'weight'
    if n.startswith('destination'):
        return 'destination'
    return None


WASTE_DATA_SHEET = 'Waste Data'


def _records_from_matrix(header_row, data_rows) -> List[Dict[str, Any]]:
    """
    Shared logic: given a header row + iterable of data rows (each a sequence of cell
    values), map columns to canonical fields and emit row dicts. Fully-empty rows are
    skipped; `row_index` is the 1-based source row number (header = row 1).
    Raises ValueError if the required Date / Waste Type headers are absent.
    """
    col_map: Dict[int, str] = {}
    for idx, cell in enumerate(header_row or []):
        canon = _canonical_header(cell) if cell is not None else None
        if canon and canon not in col_map.values():
            col_map[idx] = canon
    if 'date' not in col_map.values() or 'waste_type' not in col_map.values():
        raise ValueError('File is missing required headers (Date / Waste Type)')

    out: List[Dict[str, Any]] = []
    row_number = 1  # header was row 1
    for row in data_rows:
        row_number += 1
        if row is None or all(c is None or _clean(c) == '' for c in row):
            continue  # skip fully-empty rows
        record: Dict[str, Any] = {'row_index': row_number}
        for idx, field in col_map.items():
            record[field] = row[idx] if idx < len(row) else None
        out.append(record)
    return out


def _looks_like_xlsx(file_bytes: bytes, filename: Optional[str]) -> bool:
    """xlsx files are ZIP archives (magic 'PK\\x03\\x04'). Fall back to the extension."""
    if file_bytes[:2] == b'PK':
        return True
    if filename:
        low = filename.lower()
        if low.endswith('.csv'):
            return False
        if low.endswith(('.xlsx', '.xlsm', '.xls')):
            return True
    return False


def _parse_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse the **Waste Data** sheet only (any other tab, e.g. 'Org Chart', is ignored)."""
    import openpyxl  # local import: keeps module import cheap when parsing isn't needed

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    # Prefer the exact sheet name; fall back to a case-insensitive match; else the first sheet.
    sheet_name = None
    for name in wb.sheetnames:
        if name == WASTE_DATA_SHEET:
            sheet_name = name
            break
    if sheet_name is None:
        for name in wb.sheetnames:
            if _normalize(name) == _normalize(WASTE_DATA_SHEET):
                sheet_name = name
                break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0] if wb.sheetnames else None
    if sheet_name is None:
        raise ValueError('Workbook has no sheets')

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError('Sheet is empty')
    records = _records_from_matrix(header_row, rows_iter)
    wb.close()
    return records


def _decode_csv_bytes(file_bytes: bytes) -> str:
    """Decode CSV bytes, tolerating a UTF-8 BOM and Thai encodings (utf-8 → cp874 → latin-1)."""
    for enc in ('utf-8-sig', 'utf-8', 'cp874', 'tis-620', 'latin-1'):
        try:
            return file_bytes.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_bytes.decode('utf-8', errors='replace')


def _parse_csv(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse a single-table CSV (a CSV has no sheets — the whole file is the data)."""
    import csv

    text = _decode_csv_bytes(file_bytes)
    # Sniff the delimiter (comma/semicolon/tab); default to comma.
    sample = text[:4096]
    delimiter = ','
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=',;\t').delimiter
    except Exception:
        delimiter = ','
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise ValueError('CSV file is empty')
    return _records_from_matrix(all_rows[0], iter(all_rows[1:]))


def parse_waste_data(file_bytes: bytes, filename: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Parse an uploaded .xlsx **or** .csv into a list of raw row dicts.

    For xlsx, only the **Waste Data** sheet is read (other tabs are ignored). For csv the
    whole file is one table. Format is detected by magic bytes (xlsx = ZIP 'PK') with the
    filename extension as a fallback. Fully-empty rows are skipped. Each returned dict carries
    the 1-based source row number as `row_index` plus the canonical columns:
        date, level1..level4, tag, tenant, waste_type, weight (all raw cell values).
    An OPTIONAL trailing `destination` column (header "Destination") is also picked up when
    present (9 canonical columns without it, 10 with) — mapped by header name, so column
    order stays flexible. Raises ValueError if the required Date / Waste Type headers are missing.
    """
    if _looks_like_xlsx(file_bytes, filename):
        return _parse_xlsx(file_bytes)
    return _parse_csv(file_bytes)
