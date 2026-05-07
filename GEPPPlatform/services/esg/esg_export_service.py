"""
ESG Export Service — Generate Excel and PDF exports of collected data
"""

import io
import uuid
from datetime import datetime

import boto3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from GEPPPlatform.models.esg.records import EsgRecord


class EsgExportService:

    def __init__(self, session, s3_bucket: str = None):
        self.session = session
        self.s3_bucket = s3_bucket or 'gepp-esg-exports'
        self.s3_client = boto3.client('s3')

    def export_to_excel(self, organization_id: int) -> dict:
        """Generate .xlsx and return download link."""
        entries = self._query_entries(organization_id)
        wb = self._create_excel_workbook(entries)
        return self._upload_workbook(wb, organization_id, 'xlsx')

    def export_to_pdf(self, organization_id: int) -> dict:
        """Generate PDF report and return download link."""
        entries = self._query_entries(organization_id)
        pdf_bytes = self._create_pdf(entries, organization_id)
        return self._upload_bytes(pdf_bytes, organization_id, 'pdf', 'application/pdf')

    def _query_entries(self, organization_id: int) -> list:
        return (
            self.session.query(EsgRecord)
            .filter(
                EsgRecord.organization_id == organization_id,
                EsgRecord.is_active == True,
            )
            .order_by(EsgRecord.entry_date.desc())
            .all()
        )

    def _create_excel_workbook(self, entries: list) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = 'ESG Data'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='76B900', end_color='76B900', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = ['#', 'Source', 'Category', 'Value', 'Unit', 'tCO2e', 'Date', 'Scope', 'Status', 'Evidence', 'Notes']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, entry in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=entry.entry_source.value if entry.entry_source else '')
            ws.cell(row=row_idx, column=3, value=entry.category or str(entry.category_id or ''))
            ws.cell(row=row_idx, column=4, value=float(entry.value) if entry.value else 0)
            ws.cell(row=row_idx, column=5, value=entry.unit or '')
            ws.cell(row=row_idx, column=6, value=float(entry.calculated_tco2e) if entry.calculated_tco2e else '')
            ws.cell(row=row_idx, column=7, value=str(entry.entry_date) if entry.entry_date else '')
            ws.cell(row=row_idx, column=8, value=entry.scope_tag or '')
            ws.cell(row=row_idx, column=9, value=entry.status.value if entry.status else '')
            ws.cell(row=row_idx, column=10, value=entry.file_name or '')
            ws.cell(row=row_idx, column=11, value=entry.notes or '')

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        return wb

    def _create_pdf(self, entries: list, organization_id: int) -> bytes:
        """Generate a simple PDF report using reportlab."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('GEPP ESG Data Report', styles['Title']))
        elements.append(Paragraph(f'Organization ID: {organization_id} | Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        total_tco2e = sum(float(e.calculated_tco2e or 0) for e in entries)
        elements.append(Paragraph(f'Total tCO2e: {total_tco2e:.4f}', styles['Heading2']))
        elements.append(Paragraph(f'Total Entries: {len(entries)}', styles['Normal']))
        elements.append(Spacer(1, 20))

        # Table
        data = [['#', 'Source', 'Category', 'Value', 'Unit', 'tCO2e', 'Date', 'Scope', 'Status']]
        for i, entry in enumerate(entries[:100], 1):
            data.append([
                str(i),
                entry.entry_source.value if entry.entry_source else '',
                entry.category or '',
                f'{float(entry.value):.2f}' if entry.value else '',
                entry.unit or '',
                f'{float(entry.calculated_tco2e):.4f}' if entry.calculated_tco2e else '-',
                str(entry.entry_date) if entry.entry_date else '',
                entry.scope_tag or '',
                entry.status.value if entry.status else '',
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#76B900')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(table)

        doc.build(elements)
        return buffer.getvalue()

    def _upload_workbook(self, wb: Workbook, organization_id: int, ext: str) -> dict:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return self._upload_bytes(buffer.getvalue(), organization_id, ext, content_type)

    def _upload_bytes(self, data: bytes, organization_id: int, ext: str, content_type: str) -> dict:
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        file_key = f'exports/org_{organization_id}/esg_report_{timestamp}_{uuid.uuid4().hex[:8]}.{ext}'
        file_name = f'esg_report_{timestamp}.{ext}'

        self.s3_client.put_object(
            Bucket=self.s3_bucket, Key=file_key,
            Body=data, ContentType=content_type,
        )

        download_url = self.s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': self.s3_bucket, 'Key': file_key},
            ExpiresIn=3600,
        )

        # Nest the payload under `data` to match the rest of the ESG API
        # surface — the frontend's normalizeResponse picks `data.data` when a
        # `success` key is present, so a flat shape would arrive as undefined
        # and trip the "ไม่ได้รับลิงก์ดาวน์โหลดจากระบบ" error in LiffExcelDownload.
        return {
            'success': True,
            'data': {
                'download_url': download_url,
                'file_name': file_name,
                'format': ext,
                'expires_in': 3600,
            },
        }
