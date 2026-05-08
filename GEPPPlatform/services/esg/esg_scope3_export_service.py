"""
ESG Scope 3 Export Service — TGO CFO audit-grade workbook generator.

Reads `assets/scope3_export_template.yaml` (the canonical layout
extracted from the official T-Ver template) and writes a workbook that
mirrors that layout, populated entirely from the org's own
`esg_records` rows. The output is intended to be the customer-facing
artefact for TGO CFO Scope 3 certificate submission, so format parity
matters: every header phrase, every column order, every section
divider matches the source.

Two output modes:
  • Full          — summary sheet + 15 detail sheets.
  • Mini (cat=N)  — summary sheet (one cat row populated) + that cat's
                    single detail sheet. Same workbook structure so the
                    file drops straight into the TGO template if needed.

Customer data NEVER comes from the template file. The template is
layout-only.
"""

from __future__ import annotations

import io
import json
import logging
import os
import uuid
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from functools import lru_cache
from typing import Any, Optional

import boto3
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import extract
from sqlalchemy.orm import Session

from GEPPPlatform.models.esg.records import EsgRecord

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────
# Template loader
# ──────────────────────────────────────────────────────────────────

_ASSETS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    'assets',
)
_TEMPLATE_PATH = os.path.join(_ASSETS_DIR, 'scope3_export_template.yaml')


@lru_cache(maxsize=1)
def _load_template() -> dict:
    with open(_TEMPLATE_PATH, encoding='utf-8') as f:
        return yaml.safe_load(f)


# ──────────────────────────────────────────────────────────────────
# Deep-link helper — Excel rows hyperlink straight to the platform's
# Data Warehouse, opening the matching Scope 3 modal + evidence
# sub-modal. The link is gated by the platform's auth wall:
# ProtectedRoute redirects to /login?redirect=<encoded> if the user
# isn't signed in, and Login.tsx restores the deep link after sign-in.
# ──────────────────────────────────────────────────────────────────


def _platform_base_url() -> str:
    """Origin of the desktop ESG web app (no LIFF — we want the auth
    wall + the full data-warehouse UI)."""
    base = (
        os.getenv('ESG_WEB_BASE_URL')
        or os.getenv('LIFF_BASE_URL')
        or 'https://esg.gepp.me'
    ).rstrip('/')
    return base


def _record_deeplink(scope3_category_id: int, record_id: int) -> str:
    """Build a URL that opens the Data Warehouse, pops the matching
    Scope 3 records modal, and auto-opens that record's evidence
    sub-modal."""
    from urllib.parse import urlencode
    qs = urlencode({
        'openCat': scope3_category_id,
        'openRecord': record_id,
    })
    return f'{_platform_base_url()}/data-warehouse?{qs}'


# Visual treatment for hyperlink cells: blue + underline, matching the
# default Excel "Hyperlink" cell style so auditors see a familiar cue.
_HYPERLINK_FONT_KW = {
    'color': '0563C1',
    'underline': 'single',
}


# ──────────────────────────────────────────────────────────────────
# Helpers — datapoint resolution
# ──────────────────────────────────────────────────────────────────

# Mirrors the modal's field-name humaniser so a record's JSONB
# `datapoints` array can be looked up by canonical key from the YAML
# (e.g. `datapoint:disposal_method` → search for "disposal method"
# / "disposal_method" / "วิธีกำจัด").
_DP_ALIASES = {
    'disposal_method': ['disposal method', 'disposal', 'วิธีกำจัด'],
    'supplier':        ['supplier', 'vendor', 'ผู้ขาย'],
    'transport_mode':  ['transport mode', 'mode', 'รูปแบบขนส่ง'],
    'weight_kg':       ['weight', 'weight (kg)', 'cargo weight', 'น้ำหนัก'],
    'distance_km':     ['distance', 'distance (km)', 'travel distance', 'ระยะทาง'],
    'nights':          ['nights', 'จำนวนคืน', 'room nights'],
    'asset_type':      ['asset type', 'type', 'ประเภท'],
    'purchase_date':   ['purchase date', 'date', 'วันที่ซื้อ'],
    'quantity':        ['quantity', 'qty', 'จำนวน'],
    'unit_cost':       ['unit cost', 'unit price', 'price per unit', 'ราคาต่อหน่วย'],
    'total_cost':      ['total cost', 'amount', 'ราคารวม', 'มูลค่ารวม'],
    'material_group':  ['material group'],
    'plant':           ['plant'],
    'vendor':          ['vendor', 'supplier', 'supplier/supplying plant', 'ผู้ขาย'],
}


def _find_dp(datapoints: list, key: str) -> Optional[dict]:
    """Best-effort match of a datapoint by canonical key."""
    if not datapoints:
        return None
    aliases = set(_DP_ALIASES.get(key, [key]))
    aliases.add(key)
    aliases.add(key.replace('_', ' '))
    targets = {a.lower().strip() for a in aliases}
    for d in datapoints:
        for field in ('datapoint_name', 'canonical_name'):
            v = (d.get(field) or '').strip().lower()
            if v and v in targets:
                return d
        for tag in d.get('tags') or []:
            if isinstance(tag, str) and tag.strip().lower() in targets:
                return d
    return None


def _dp_value(datapoints: list, key: str) -> Any:
    d = _find_dp(datapoints, key)
    if d is None:
        return None
    return d.get('value')


def _dp_sum(datapoints_list: list, key: str) -> float:
    """Sum a datapoint's numeric value across many records."""
    total = 0.0
    for dps in datapoints_list:
        d = _find_dp(dps or [], key)
        if d is None:
            continue
        try:
            total += float(d.get('value'))
        except (TypeError, ValueError):
            pass
    return total


# ──────────────────────────────────────────────────────────────────
# Service
# ──────────────────────────────────────────────────────────────────


class EsgScope3ExportService:
    """Build the TGO-grade Scope 3 workbook for an organisation."""

    def __init__(self, session: Session, s3_bucket: Optional[str] = None):
        self.session = session
        self.s3_bucket = s3_bucket or os.getenv(
            'S3_BUCKET_NAME', 'gepp-esg-exports',
        )
        self.s3_client = boto3.client('s3')
        self.tpl = _load_template()

    # ── public ──

    def export(
        self,
        organization_id: int,
        year: int,
        scope3_category_id: Optional[int] = None,
        push_to_line_user_id: Optional[str] = None,
    ) -> dict:
        records_by_cat = self._fetch_records(organization_id, year)
        wb = self._build_workbook(
            records_by_cat=records_by_cat,
            year=year,
            target_cat=scope3_category_id,
        )
        result = self._upload_workbook(wb, organization_id, year, scope3_category_id)

        # Optional LINE-chat delivery. The xlsx itself can't be sent
        # through the Push API (no `file` content type), so we send
        # a Flex card carrying a button → presigned download URL.
        # The user lands in LINE chat where Android's external-browser
        # "Open in" prompt works reliably.
        if push_to_line_user_id:
            try:
                payload = result.get('data') or {}
                pushed = self._push_to_line(
                    line_user_id=push_to_line_user_id,
                    download_url=payload.get('download_url'),
                    file_name=payload.get('file_name'),
                    year=year,
                    scope3_category_id=scope3_category_id,
                )
                payload['pushed'] = bool(pushed)
                payload['delivery'] = 'line_chat' if pushed else 'browser'
                result['data'] = payload
            except Exception as e:
                logger.exception('Scope3 export — LINE push failed')
                payload = result.get('data') or {}
                payload['pushed'] = False
                payload['push_error'] = str(e)
                result['data'] = payload
        return result

    # ── data fetch ──

    def _fetch_records(
        self, organization_id: int, year: int,
    ) -> dict[int, list[EsgRecord]]:
        rows = (
            self.session.query(EsgRecord)
            .filter(
                EsgRecord.organization_id == organization_id,
                EsgRecord.is_active == True,
                extract('year', EsgRecord.entry_date) == year,
                EsgRecord.scope3_category_id.isnot(None),
            )
            .order_by(EsgRecord.scope3_category_id, EsgRecord.entry_date)
            .all()
        )
        out: dict[int, list[EsgRecord]] = defaultdict(list)
        for r in rows:
            out[int(r.scope3_category_id)].append(r)
        return out

    # ── workbook ──

    def _build_workbook(
        self,
        records_by_cat: dict[int, list[EsgRecord]],
        year: int,
        target_cat: Optional[int] = None,
    ) -> Workbook:
        wb = Workbook()
        # Drop the default sheet — we add named sheets explicitly.
        del wb[wb.sheetnames[0]]

        # Always write the summary sheet first so an auditor sees totals.
        self._write_summary_sheet(wb, records_by_cat, year, target_cat)

        # Detail sheets — full set in full mode, single sheet in mini mode.
        cats_to_write = (
            [c for c in self.tpl['categories'] if c['id'] == target_cat]
            if target_cat
            else list(self.tpl['categories'])
        )
        for cat_def in cats_to_write:
            cid = cat_def['id']
            self._write_detail_sheet(
                wb=wb,
                cat_def=cat_def,
                records=records_by_cat.get(cid) or [],
                year=year,
            )
        return wb

    # ── summary ──

    def _write_summary_sheet(
        self,
        wb: Workbook,
        records_by_cat: dict[int, list[EsgRecord]],
        year: int,
        target_cat: Optional[int],
    ) -> None:
        s = self.tpl['summary']
        ws = wb.create_sheet(s['sheet_name'])

        labels = self.tpl['category_labels']
        title_font = Font(bold=True, size=self.tpl['style']['title_size'])
        header_font = Font(
            bold=True,
            size=self.tpl['style']['header_size'],
            color=self.tpl['style']['header_font_color'],
        )
        header_fill = PatternFill(
            start_color=self.tpl['style']['header_fill'],
            end_color=self.tpl['style']['header_fill'],
            fill_type='solid',
        )
        thin = Side(style='thin', color=self.tpl['style']['border_color'])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title
        ws.cell(row=1, column=1, value=f'Scope 3 — Summary (Year {year})').font = title_font

        # Header row
        for col in s['columns']:
            cell = ws.cell(row=s['header_row'], column=col['col'])
            cell.value = f"{col['label_th']} / {col['label_en']}"
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col['col'])].width = col.get('width', 18)

        # Per-cat totals
        totals_t = {}  # cat → tCO2e
        for cid in range(1, 16):
            kg_sum = 0.0
            for r in records_by_cat.get(cid) or []:
                if r.kgco2e is not None:
                    try:
                        kg_sum += float(r.kgco2e)
                    except (TypeError, ValueError):
                        pass
            totals_t[cid] = kg_sum / 1000.0

        grand_total = sum(totals_t.values()) or 0.0

        row = s['data_start_row']
        for cid in range(1, 16):
            if target_cat and cid != target_cat:
                # Mini mode — write only the target row, skip the rest
                continue
            label = labels.get(cid, {})
            ws.cell(row=row, column=1, value=cid).border = border
            cell_label = ws.cell(row=row, column=2, value=label.get('th') or label.get('en') or '')
            cell_label.border = border
            cell_label.alignment = Alignment(wrap_text=True)
            cell_t = ws.cell(row=row, column=3, value=round(totals_t[cid], 4))
            cell_t.border = border
            cell_t.number_format = '#,##0.0000'
            pct = (totals_t[cid] / grand_total) if grand_total > 0 else 0
            cell_p = ws.cell(row=row, column=4, value=pct)
            cell_p.border = border
            cell_p.number_format = '0.00%'
            row += 1

        # Grand total row
        if not target_cat:
            ws.cell(row=row, column=2, value='Total').font = Font(bold=True)
            cell_gt = ws.cell(row=row, column=3, value=round(grand_total, 4))
            cell_gt.font = Font(bold=True)
            cell_gt.number_format = '#,##0.0000'
            cell_gp = ws.cell(row=row, column=4, value=1 if grand_total > 0 else 0)
            cell_gp.font = Font(bold=True)
            cell_gp.number_format = '0.00%'

    # ── detail sheets ──

    def _write_detail_sheet(
        self,
        wb: Workbook,
        cat_def: dict,
        records: list[EsgRecord],
        year: int,
    ) -> None:
        ws = wb.create_sheet(cat_def['sheet_name'])
        self._write_sheet_header(ws, cat_def, year)

        kind = cat_def.get('kind') or 'monthly'
        # Placeholder cats with no activity: render the "ไม่มีกิจกรรม" stub
        if cat_def.get('placeholder') and not records:
            self._write_no_activity(ws)
            return

        cat_id = int(cat_def.get('id') or 0)
        if kind == 'monthly':
            self._write_monthly_table(ws, cat_def, records, year, cat_id)
        elif kind == 'asset_register':
            self._write_flat_table(ws, cat_def, records, cat_id, with_seq=True)
        elif kind == 'travel_detail':
            self._write_flat_table(ws, cat_def, records, cat_id, with_seq=True)
        elif kind == 'headcount_aggregate':
            self._write_aggregate_block(ws, cat_def, records)
        else:
            self._write_monthly_table(ws, cat_def, records, year, cat_id)

    def _write_sheet_header(self, ws, cat_def: dict, year: int) -> None:
        h = self.tpl['sheet_header']
        title = h['title_row']
        ws.cell(
            row=title['row'],
            column=title['col'],
            value=title['value_th'],
        ).font = Font(bold=True, size=self.tpl['style']['title_size'])
        # Topic
        ws.cell(row=h['topic_row']['row'], column=h['topic_row']['key_col'],
                value=h['topic_row']['key_th']).font = Font(bold=True)
        ws.cell(row=h['topic_row']['row'], column=h['topic_row']['value_col'],
                value=cat_def['title_th'])
        # Period
        months_th = self.tpl['months_th']
        period_value = h['period_row']['value_format'].format(
            month_start_th=months_th[0],
            month_end_th=months_th[11],
            year=year,
        )
        ws.cell(row=h['period_row']['row'], column=h['period_row']['key_col'],
                value=h['period_row']['key_th']).font = Font(bold=True)
        ws.cell(row=h['period_row']['row'], column=h['period_row']['value_col'],
                value=period_value)

    def _apply_record_hyperlink(
        self,
        cell,
        scope3_category_id: int,
        record_id: int,
    ) -> None:
        """Turn `cell` into a hyperlink that opens the platform's Data
        Warehouse on the right Scope 3 cat with the evidence sub-modal
        for `record_id` already open. The cell's existing font is
        replaced with the blue/underline treatment Excel uses for
        Hyperlink-style cells."""
        if not record_id or not scope3_category_id:
            return
        cell.hyperlink = _record_deeplink(scope3_category_id, record_id)
        existing = cell.font
        # `Font.copy(...)` is removed in newer openpyxl — rebuild instead.
        cell.font = Font(
            name=existing.name,
            size=existing.size,
            bold=existing.bold,
            italic=existing.italic,
            color=_HYPERLINK_FONT_KW['color'],
            underline=_HYPERLINK_FONT_KW['underline'],
        )

    def _write_no_activity(self, ws) -> None:
        c = self.tpl['no_activity']['cell']
        ws.cell(row=c['row'], column=c['col'],
                value=self.tpl['no_activity']['th']).font = Font(italic=True, color='808080')

    # ── kind = "monthly": item rows × Jan-Dec columns + Total ──

    def _write_monthly_table(
        self,
        ws,
        cat_def: dict,
        records: list[EsgRecord],
        year: int,
        scope3_category_id: int = 0,
    ) -> None:
        cols = list(cat_def.get('columns') or [])
        header_row = self.tpl['sheet_header']['table_start_row']
        data_start_row = header_row + 1

        thin = Side(style='thin', color=self.tpl['style']['border_color'])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(
            bold=True,
            size=self.tpl['style']['header_size'],
            color=self.tpl['style']['header_font_color'],
        )
        header_fill = PatternFill(
            start_color=self.tpl['style']['header_fill'],
            end_color=self.tpl['style']['header_fill'],
            fill_type='solid',
        )

        # Header — fixed columns, then 12 monthly columns, then Total
        col_idx = 1
        col_specs: list[tuple[int, dict]] = []  # (col_index, col_def)
        for c in cols:
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = f"{c['label_th']} / {c['label_en']}"
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = c.get('width', 16)
            col_specs.append((col_idx, c))
            col_idx += 1

        # Monthly columns
        months_th = self.tpl['months_th']
        month_first_col = col_idx
        for m in range(12):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = f"{months_th[m]} {year}"
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            col_idx += 1

        # Total + Notes
        total_col = None
        if cat_def.get('has_total'):
            total_col = col_idx
            cell = ws.cell(row=header_row, column=col_idx, value='Total / รวม')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            col_idx += 1
        if cat_def.get('has_notes'):
            cell = ws.cell(row=header_row, column=col_idx, value='หมายเหตุ / Notes')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = 28
            col_idx += 1

        # Group records by record_label (one item = one row)
        groups: dict[str, list[EsgRecord]] = defaultdict(list)
        for r in records:
            label = (r.record_label or f'Record #{r.id}').strip() or f'Record #{r.id}'
            groups[label].append(r)

        if not groups:
            self._write_no_activity(ws)
            return

        zebra = PatternFill(
            start_color=self.tpl['style']['zebra_fill'],
            end_color=self.tpl['style']['zebra_fill'],
            fill_type='solid',
        )

        row = data_start_row
        for seq_idx, (label, group_records) in enumerate(groups.items(), start=1):
            sample = group_records[0]
            datapoints = list(sample.datapoints or [])
            link_record_id = int(getattr(sample, 'id', 0) or 0)

            # Fill fixed columns
            for col_index, col_def in col_specs:
                value = self._resolve_value(
                    col_def, sample, datapoints, seq_idx, label,
                )
                cell = ws.cell(row=row, column=col_index, value=value)
                cell.border = border
                if col_def.get('number_format'):
                    cell.number_format = col_def['number_format']
                if col_index % 2 == 0:
                    cell.fill = zebra
                # Make the row's identifying cell (record label / item)
                # a hyperlink back to the platform's Data Warehouse with
                # the evidence sub-modal pre-opened for this record.
                if col_def.get('key') == 'record_label':
                    self._apply_record_hyperlink(
                        cell, scope3_category_id, link_record_id,
                    )

            # Monthly: pivot kgco2e (or quantity if available) per month
            month_totals = [0.0] * 12
            for r in group_records:
                if not r.entry_date:
                    continue
                m = r.entry_date.month - 1
                if 0 <= m <= 11:
                    qty = self._record_quantity(r)
                    month_totals[m] += qty

            row_total = 0.0
            for m in range(12):
                v = month_totals[m]
                row_total += v
                if v:
                    cell = ws.cell(row=row, column=month_first_col + m, value=round(v, 4))
                else:
                    cell = ws.cell(row=row, column=month_first_col + m, value=0)
                cell.number_format = '#,##0.00'
                cell.border = border

            if total_col is not None:
                cell = ws.cell(row=row, column=total_col, value=round(row_total, 4))
                cell.number_format = '#,##0.00'
                cell.border = border
                cell.font = Font(bold=True)

            row += 1

    # ── kind = "asset_register" / "travel_detail": flat list ──

    def _write_flat_table(
        self,
        ws,
        cat_def: dict,
        records: list[EsgRecord],
        scope3_category_id: int = 0,
        with_seq: bool = True,
    ) -> None:
        cols = list(cat_def.get('columns') or [])
        header_row = self.tpl['sheet_header']['table_start_row']
        data_start_row = header_row + 1

        thin = Side(style='thin', color=self.tpl['style']['border_color'])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(
            bold=True,
            size=self.tpl['style']['header_size'],
            color=self.tpl['style']['header_font_color'],
        )
        header_fill = PatternFill(
            start_color=self.tpl['style']['header_fill'],
            end_color=self.tpl['style']['header_fill'],
            fill_type='solid',
        )

        # Header
        for idx, c in enumerate(cols, start=1):
            cell = ws.cell(row=header_row, column=idx)
            cell.value = f"{c['label_th']} / {c['label_en']}"
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(idx)].width = c.get('width', 16)

        if not records:
            self._write_no_activity(ws)
            return

        for seq_idx, r in enumerate(records, start=1):
            datapoints = list(r.datapoints or [])
            link_record_id = int(getattr(r, 'id', 0) or 0)
            for col_index, col_def in enumerate(cols, start=1):
                value = self._resolve_value(
                    col_def, r, datapoints, seq_idx, r.record_label or '',
                )
                cell = ws.cell(row=data_start_row + seq_idx - 1, column=col_index, value=value)
                cell.border = border
                if col_def.get('number_format'):
                    cell.number_format = col_def['number_format']
                if col_def.get('key') == 'record_label':
                    self._apply_record_hyperlink(
                        cell, scope3_category_id, link_record_id,
                    )

    # ── kind = "headcount_aggregate" (Cat 7) ──

    def _write_aggregate_block(
        self,
        ws,
        cat_def: dict,
        records: list[EsgRecord],
    ) -> None:
        rows = list(cat_def.get('aggregate_rows') or [])
        header_row = self.tpl['sheet_header']['table_start_row']

        thin = Side(style='thin', color=self.tpl['style']['border_color'])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(
            bold=True,
            size=self.tpl['style']['header_size'],
            color=self.tpl['style']['header_font_color'],
        )

        # Headers
        ws.cell(row=header_row, column=2, value='Metric').font = header_font
        ws.cell(row=header_row, column=3, value='Value').font = header_font
        ws.cell(row=header_row, column=4, value='Unit').font = header_font

        if not records:
            self._write_no_activity(ws)
            return

        all_dps = [r.datapoints or [] for r in records]
        kg_total = sum(
            (float(r.kgco2e) if r.kgco2e is not None else 0.0)
            for r in records
        )

        out_row = header_row + 1
        for spec in rows:
            ws.cell(row=out_row, column=2, value=f"{spec['label_th']} / {spec['label_en']}").border = border
            key = spec.get('key') or ''
            val: Any = ''
            if key.startswith('datapoint_sum:'):
                val = round(_dp_sum(all_dps, key.split(':', 1)[1]), 4)
            elif key == 'kgco2e_year_total':
                val = round(kg_total, 4)
            elif key == 'metric:car_headcount':
                val = len({r.line_user_id or r.user_id for r in records if (r.line_user_id or r.user_id)})
            elif key == 'metric:fuel_estimate_litres':
                # ~14.763 km/L benchmark for petrol — matches the T-Ver
                # template's footnote.
                km = _dp_sum(all_dps, 'distance_km')
                val = round(km / 14.763, 4) if km else 0
            else:
                val = ''
            cell = ws.cell(row=out_row, column=3, value=val)
            cell.number_format = '#,##0.0000'
            cell.border = border
            ws.cell(row=out_row, column=4, value=spec.get('unit') or '').border = border
            out_row += 1

    # ── value resolver ──

    def _resolve_value(
        self,
        col_def: dict,
        sample: EsgRecord,
        datapoints: list,
        seq_idx: int,
        label: str,
    ) -> Any:
        key = col_def.get('key') or ''
        if not key:
            return ''
        if key == 'seq':
            return seq_idx
        if key == 'record_label':
            return label or sample.record_label or ''
        if key == 'unit':
            # Pull the unit off the most-numeric datapoint
            for d in datapoints:
                u = (d.get('unit') or '').strip()
                if u:
                    return u
            return ''
        if key == 'kgco2e':
            return float(sample.kgco2e) if sample.kgco2e is not None else ''
        if key == 'evidence_ref':
            return sample.evidence_image_url or sample.file_key or ''
        if key == 'responsible':
            # We don't store responsible-person yet; surface line_user_id
            # so the auditor at least has *some* attribution.
            return sample.line_user_id or ''
        if key.startswith('datapoint:'):
            v = _dp_value(datapoints, key.split(':', 1)[1])
            if isinstance(v, (Decimal,)):
                return float(v)
            return v if v is not None else ''
        return ''

    # ── helper: pick the "quantity" of a record for the monthly pivot ──

    def _record_quantity(self, r: EsgRecord) -> float:
        """
        Pick the quantitative measurement to plot in the monthly cell.
        Order:
          1. kgco2e (always defined for `computed` records — most useful
             for an auditor). Falls through to:
          2. The first datapoint that has a numeric value with a
             physical unit (kg, km, kWh, …).
          3. 0.
        """
        if r.kgco2e is not None:
            try:
                return float(r.kgco2e)
            except (TypeError, ValueError):
                pass
        for d in (r.datapoints or []):
            v = d.get('value')
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            if fv:
                return fv
        return 0.0

    # ── LINE push ──

    def _push_to_line(
        self,
        line_user_id: str,
        download_url: Optional[str],
        file_name: Optional[str],
        year: int,
        scope3_category_id: Optional[int],
    ) -> bool:
        """
        Push a Flex message to the LINE user with a button → presigned
        download URL. This is the workaround for LINE's Android
        in-app-browser blocking direct file responses.

        Returns True on push success.
        """
        if not line_user_id or not download_url:
            return False
        token = (
            os.getenv('ESG_LINE_CHANNEL_ACCESS_TOKEN')
            or os.getenv('LINE_CHANNEL_ACCESS_TOKEN')
        )
        if not token:
            logger.warning(
                'Scope3 push: ESG_LINE_CHANNEL_ACCESS_TOKEN missing; cannot push',
            )
            return False

        # Route through the LIFF entry URL so the link opens inside the
        # LIFF in-app browser (consistent with all other LINE Card
        # links in the product). The LIFF prefix already represents
        # `/liff` on the host, so appending `/download` resolves to
        # `/liff/download` — the DownloadLanding route — which then
        # breaks out to the external browser via
        # liff.openWindow({ external: true }) for the actual file
        # download (Android's LINE in-app browser blocks direct
        # Content-Disposition: attachment responses).
        liff_entry = os.getenv(
            'LIFF_ENTRY_BASE',
            'https://liff.line.me/2009993849-GpYCVVmc',
        ).rstrip('/')
        from urllib.parse import urlencode
        landing_url = (
            f'{liff_entry}/download?'
            + urlencode({
                'url': download_url,
                'name': file_name or 'gepp-esg-scope3.xlsx',
            })
        )

        cat_label = (
            f'Cat {scope3_category_id}'
            if scope3_category_id else 'ทั้ง 15 categories'
        )
        title_th = f'รายงาน Scope 3 ปี {year}'
        subtitle_th = (
            f'รูปแบบ TGO CFO · {cat_label}\n'
            'กดปุ่มด้านล่างเพื่อเปิดในเบราว์เซอร์และดาวน์โหลด'
        )

        flex = {
            'type': 'flex',
            'altText': f'{title_th} ({cat_label}) — แตะเพื่อดาวน์โหลด',
            'contents': {
                'type': 'bubble',
                'size': 'kilo',
                'body': {
                    'type': 'box', 'layout': 'vertical', 'spacing': 'sm',
                    'paddingAll': '20px',
                    'contents': [
                        {
                            'type': 'text', 'text': 'TGO CFO · SCOPE 3',
                            'size': 'xs', 'color': '#10b981',
                            'weight': 'bold',
                        },
                        {
                            'type': 'text', 'text': title_th,
                            'size': 'xl', 'weight': 'bold',
                            'wrap': True, 'color': '#0b1120',
                            'margin': 'sm',
                        },
                        {
                            'type': 'text', 'text': subtitle_th,
                            'size': 'sm', 'wrap': True,
                            'color': '#475569', 'margin': 'md',
                        },
                        {
                            'type': 'box', 'layout': 'vertical',
                            'backgroundColor': '#f1f5f9',
                            'cornerRadius': '8px',
                            'paddingAll': '10px',
                            'margin': 'md',
                            'contents': [{
                                'type': 'text',
                                'text': f'📊  {file_name or "scope3.xlsx"}',
                                'size': 'xs', 'color': '#475569',
                                'wrap': True,
                            }],
                        },
                    ],
                },
                'footer': {
                    'type': 'box', 'layout': 'vertical', 'spacing': 'sm',
                    'paddingAll': '14px',
                    'contents': [{
                        'type': 'button',
                        'style': 'primary',
                        'color': '#0B1120',
                        'action': {
                            'type': 'uri',
                            'label': 'ดาวน์โหลด Excel',
                            'uri': landing_url,
                        },
                    }],
                },
            },
        }

        try:
            import json as _json
            import urllib.request
            req = urllib.request.Request(
                'https://api.line.me/v2/bot/message/push',
                data=_json.dumps(
                    {'to': line_user_id, 'messages': [flex]},
                ).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'Authorization': f'Bearer {token}',
                },
            )
            with urllib.request.urlopen(req, timeout=10) as resp:
                status = getattr(resp, 'status', None) or resp.getcode()
                logger.info(
                    'Scope3 LINE push OK status=%s to=%s file=%s',
                    status, line_user_id[:12], file_name,
                )
                return True
        except Exception:
            logger.exception('Scope3 LINE push failed')
            return False

    # ── upload + presign ──

    def _upload_workbook(
        self,
        wb: Workbook,
        organization_id: int,
        year: int,
        scope3_category_id: Optional[int],
    ) -> dict:
        buffer = io.BytesIO()
        wb.save(buffer)
        data = buffer.getvalue()
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        cat_part = f'-cat{scope3_category_id}' if scope3_category_id else ''
        file_key = (
            f'exports/org_{organization_id}/scope3-{year}{cat_part}'
            f'_{timestamp}_{uuid.uuid4().hex[:8]}.xlsx'
        )
        file_name = f'gepp-esg-scope3-{year}{cat_part}.xlsx'
        content_type = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        try:
            self.s3_client.put_object(
                Bucket=self.s3_bucket,
                Key=file_key,
                Body=data,
                ContentType=content_type,
            )
            download_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.s3_bucket, 'Key': file_key},
                ExpiresIn=3600,
            )
        except Exception:
            logger.exception('Failed to upload scope3 export to S3')
            raise

        return {
            'success': True,
            'data': {
                'download_url': download_url,
                'file_name': file_name,
                'format': 'xlsx',
                'expires_in': 3600,
                'year': year,
                'scope3_category_id': scope3_category_id,
            },
        }
